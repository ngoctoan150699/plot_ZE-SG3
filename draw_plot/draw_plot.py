#!/usr/bin/env python3
"""
CSV Torque Plot Viewer
======================
- Import CSV file with format CTR DATA or two-column (time,torque)
- Vẽ plot Time vs Torque
- Nhập STT bắt đầu/kết thúc để tính trung bình Torque
"""

import sys
import io
import os
import tempfile
from datetime import datetime
import json
import unicodedata

# Add python directory to sys.path to allow imports when running draw_plot.py standalone
_script_dir = os.path.dirname(os.path.abspath(__file__))
_project_dir = os.path.dirname(_script_dir)
_python_dir = os.path.join(_project_dir, 'python')
if _python_dir not in sys.path:
    sys.path.append(_python_dir)

# Import domain constants (resilient to standalone execution)
FALLBACK_TEAMS = ['PM', 'QM', 'PT', 'EDTV']
FALLBACK_LINE_NOS = [
    'ITR #1', 'ITR #2', 'ITR #3',
    'B/Joint #1', 'B/Joint #2', 'B/Joint #3',
    'OTR #1', 'OTR #2', 'OTR #3',
    'S/Link #1', 'S/Link #2', 'S/Link #3', 'S/Link #4',
    'Other'
]
try:
    from domain.constants import TEAMS, LINE_NOS
except Exception:
    TEAMS = FALLBACK_TEAMS
    LINE_NOS = FALLBACK_LINE_NOS

def remove_diacritics_no_strip(text: str) -> str:
    """Helper to remove Vietnamese diacritics and uppercase without stripping spaces."""
    if not text:
        return ""
    nfd_form = unicodedata.normalize('NFD', text)
    clean_chars = [c for c in nfd_form if not unicodedata.combining(c)]
    clean_text = "".join(clean_chars)
    clean_text = clean_text.replace('đ', 'd').replace('Đ', 'D').replace('đ', 'd')
    return clean_text.upper()


# ========== Excel layout defaults (module-level for easy maintenance) ==========
# Title
TITLE_MERGE = 'A1:E2'
# Metadata block rows (left and right blocks)
META_START_ROW = 3
META_END_ROW = 8
# Graph label row (GRAPH / STT row) placed directly under metadata
GRAPH_LABEL_ROW = META_END_ROW + 1
# Merged graph area (outer merged frame) - keep below metadata and label row
GRAPH_MERGE = 'A10:H32'
GRAPH_TOP = 10
GRAPH_BOTTOM = 32
# Inner image area (image will be sized to fit inside these rows)
# Set the inner image to end well above the merged bottom so the sample
# table can be placed below the outer merged frame without overlapping.
IMAGE_TOP = 11
# Let the inner image area extend to the merged graph bottom so the image
# can fill the full merged frame if requested.
IMAGE_BOTTOM = GRAPH_BOTTOM - 2
IMAGE_ANCHOR = 'B11'  # top-left cell where image will be anchored
# --- Image export tuning constants ---
# Padding passed to `fig.tight_layout(pad=...)` to reduce whitespace around axes
IMAGE_TIGHT_PAD = 0.5
# pad_inches passed to `savefig(..., pad_inches=...)`
IMAGE_PAD_INCHES = 0
# Extra pixels to expand width/height to compensate Excel margins
IMAGE_EXPAND_PX_W = 8
IMAGE_EXPAND_PX_H = 8
# Render scale when re-rendering at higher resolution before downsampling
IMAGE_RENDER_SCALE = 2
IMAGE_MAX_RENDER_PX = 8000
# EMU units per pixel for openpyxl image anchor offsets
IMAGE_EMU_PER_PIXEL = 9525
# Fraction of a column width to offset image horizontally (0.5 = half column)
# Use 0.5 to nudge the exported image right by half of column A.
IMAGE_HALF_COL_OFFSET_FRAC = 0.5
# Sample summary table start row (below the outer merged graph area)
SAMPLE_HEADER_ROW = GRAPH_BOTTOM + 1
# Minimum end row for outer border
OUTER_MIN_END_ROW = GRAPH_BOTTOM + 1

# Cell mapping for easier maintenance
# Write/Review/Approval labels and values (top-right)
WRA_LABEL_ROW = 1
WRA_VALUE_ROW = 2
WRA_LABEL_COLS = ['F', 'G', 'H']
WRA_VALUE_COLS = ['F', 'G', 'H']

# Left info block columns (labels merged A:B, values C:D)
LEFT_LABEL_COLS = ('A', 'B')
LEFT_VALUE_COLS = ('C', 'D')

# Right info block columns (labels merged E:F, values G:H)
RIGHT_LABEL_COLS = ('E', 'F')
RIGHT_VALUE_COLS = ('G', 'H')

import json

from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QGridLayout, QGroupBox, QLabel, QPushButton, QSpinBox, QDoubleSpinBox, QFileDialog, QMessageBox,
    QLineEdit, QComboBox, QDateEdit, QSizePolicy, QSplitter, QScrollArea, QFrame, QAction, QStyle, QInputDialog,
    QTabWidget
)

from convert_may_gui import ConvertWidget

from PyQt5.QtGui import QFont, QIcon, QPixmap
import pathlib


def get_config_dir():
    """Return a writable configuration directory for the application.
    On Windows this prefers %APPDATA% or %LOCALAPPDATA%; otherwise falls
    back to the user's home directory. Creates the directory if missing.
    """
    try:
        appdata = os.getenv('APPDATA') or os.getenv('LOCALAPPDATA')
        if appdata:
            base = pathlib.Path(appdata)
        else:
            base = pathlib.Path.home()
        cfg = base / 'plc_sim'
        cfg.mkdir(parents=True, exist_ok=True)
        return cfg
    except Exception:
        # last resort: use the script directory (may be read-only when frozen)
        try:
            return pathlib.Path(getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__))))
        except Exception:
            return pathlib.Path('.')


def get_config_file(name: str):
    return get_config_dir() / name

import matplotlib.pyplot as plt
from PyQt5.QtCore import QDate, Qt, QRect, QPoint, QSize, QLocale
from PyQt5.QtWidgets import QDialog, QTableWidget, QTableWidgetItem, QHeaderView, QLayout
import bisect
import pathlib

from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas, NavigationToolbar2QT
from matplotlib.figure import Figure
from matplotlib.ticker import MaxNLocator

class FlowLayout(QLayout):
    """Layout that arranges items in a flow (left to right, wrapping to new lines)."""
    def __init__(self, parent=None, margin=0, hSpacing=6, vSpacing=6):
        super(FlowLayout, self).__init__(parent)
        if parent:
            self.setContentsMargins(margin, margin, margin, margin)
        self.m_hSpace = hSpacing
        self.m_vSpace = vSpacing
        self.itemList = []

    def addItem(self, item):
        self.itemList.append(item)

    def horizontalSpacing(self):
        return self.m_hSpace

    def verticalSpacing(self):
        return self.m_vSpace

    def expandingDirections(self):
        return Qt.Orientations(0)

    def hasHeightForWidth(self):
        return True

    def heightForWidth(self, width):
        height = self.doLayout(QRect(0, 0, width, 0), True)
        return height

    def count(self):
        return len(self.itemList)

    def itemAt(self, index):
        if 0 <= index < len(self.itemList):
            return self.itemList[index]
        return None

    def minimumSize(self):
        size = QSize()
        for item in self.itemList:
            size = size.expandedTo(item.minimumSize())
        size += QSize(2*self.contentsMargins().top(), 2*self.contentsMargins().top())
        return size

    def setGeometry(self, rect):
        super(FlowLayout, self).setGeometry(rect)
        self.doLayout(rect, False)

    def sizeHint(self):
        return self.minimumSize()

    def takeAt(self, index):
        if 0 <= index < len(self.itemList):
            return self.itemList.pop(index)
        return None

    def doLayout(self, rect, testOnly):
        left, top, right, bottom = self.getContentsMargins()
        effectiveRect = rect.adjusted(+left, +top, -right, -bottom)
        x = effectiveRect.x()
        y = effectiveRect.y()
        lineHeight = 0
        
        for item in self.itemList:
            wid = item.widget()
            spaceX = self.horizontalSpacing()
            spaceY = self.verticalSpacing()
            
            # If item has no widget (e.g. spacer), skip or handle? FlowLayout usually for widgets.
            if not wid:
                continue
                
            nextX = x + item.sizeHint().width() + spaceX
            
            if nextX - spaceX > effectiveRect.right() and lineHeight > 0:
                x = effectiveRect.x()
                y = y + lineHeight + spaceY
                nextX = x + item.sizeHint().width() + spaceX
                lineHeight = 0
            
            if not testOnly:
                item.setGeometry(QRect(QPoint(x, y), item.sizeHint()))
            
            x = nextX
            lineHeight = max(lineHeight, item.sizeHint().height())
            
        return y + lineHeight - rect.y() + bottom

class CommaDoubleSpinBox(QDoubleSpinBox):
    """
    QDoubleSpinBox that accepts both comma and dot as decimal separator,
    always displays with comma, and increments by the smallest decimal unit.
    """
    def __init__(self, parent=None):
        super().__init__(parent)
        # Force English locale so internal validation always expects dot
        self.setLocale(QLocale(QLocale.English))
        self.setDecimals(9)
        self.setSingleStep(1e-6)  # Increment by smallest unit (10^-6)
        # Ensure pressing Enter or finishing edit sets value correctly
        try:
            self.editingFinished.connect(self._on_editing_finished)
        except Exception:
            pass
        # Also connect the embedded QLineEdit's returnPressed to ensure Enter works
        try:
            le = self.lineEdit()
            if le is not None:
                le.returnPressed.connect(self._on_editing_finished)
        except Exception:
            pass

    def validate(self, text, pos):
        # Allow validation to pass if comma is used by swapping for dot
        # Also ensure dot is allowed (if standard validator expects it)
        # We normalize to dot for the check
        # Normalize comma to dot for validator check
        norm = text.replace(',', '.')
        return QDoubleSpinBox.validate(self, norm, pos)

    def valueFromText(self, text):
        # Remove suffix if present before converting
        suffix = self.suffix()
        if suffix and text.endswith(suffix):
            text = text[:-len(suffix)]
        try:
            return float(text.replace(',', '.'))
        except ValueError:
            return 0.0

    def textFromValue(self, value):
        # Display with comma, respecting current decimal precision
        # Strip trailing zeros for cleaner look
        text = f"{value:.{self.decimals()}f}".replace('.', ',')
        if ',' in text:
            text = text.rstrip('0').rstrip(',')
        return text

    def keyPressEvent(self, event):
        # Accept both comma and dot as decimal separator by inserting a dot
        # into the underlying line edit so the validator/conversion works.
        try:
            ch = event.text()
        except Exception:
            ch = ''

        # If the user types either dot or comma, insert a comma so the
        # displayed text uses comma as decimal separator.
        if ch in ('.', ','):
            le = self.lineEdit()
            if le is not None:
                s = le.text()
                sel_start = le.selectionStart()
                sel_text = le.selectedText()
                if sel_start is not None and sel_start >= 0 and sel_text:
                    # Replace selected text with comma
                    new = s[:sel_start] + ',' + s[sel_start + len(sel_text):]
                    pos = sel_start + 1
                else:
                    cur = le.cursorPosition()
                    new = s[:cur] + ',' + s[cur:]
                    pos = cur + 1
                le.setText(new)
                le.setCursorPosition(pos)
                return

        # If Enter/Return pressed, finalize the edit and parse text
        from PyQt5.QtCore import Qt
        if event.key() in (Qt.Key_Return, Qt.Key_Enter):
            try:
                self._on_editing_finished()
            except Exception:
                pass
            # Let the base class handle committing if needed
            super().keyPressEvent(event)
            return

        # Fallback to default handling for other keys
        super().keyPressEvent(event)

    def focusOutEvent(self, event):
        # When leaving the widget, normalize the text and ensure the
        # underlying numeric value is set correctly (handles cases
        # where locale or intermediate text might confuse parsing).
        try:
            le = self.lineEdit()
            if le is not None:
                txt = le.text().strip()
                if txt:
                    # Replace comma with dot for Python float parsing
                    norm = txt.replace(',', '.')
                    # Remove any grouping separators or stray spaces
                    norm = norm.replace(' ', '')
                    # If user entered just a dot or comma, ignore
                    if norm not in ('.', ','):
                        try:
                            val = float(norm)
                            # Clamp to range to avoid exceptions
                            if val < self.minimum():
                                val = self.minimum()
                            if val > self.maximum():
                                val = self.maximum()
                            # Set the cleaned numeric value which will update display
                            self.setValue(val)
                        except Exception:
                            pass
        except Exception:
            pass
        super().focusOutEvent(event)

    def _on_editing_finished(self):
        # Called when user finishes editing (presses Enter)
        try:
            le = self.lineEdit()
            if le is None:
                return
            txt = le.text().strip()
            if not txt:
                return
            norm = txt.replace(',', '.')
            norm = norm.replace(' ', '')
            if norm in ('.', ','):
                return
            try:
                val = float(norm)
            except Exception:
                return
            if val < self.minimum():
                val = self.minimum()
            if val > self.maximum():
                val = self.maximum()
            # Set value which will update display text
            self.setValue(val)
        except Exception:
            pass
class CustomToolbar(NavigationToolbar2QT):
    """Custom Toolbar that removes Subplots button and adds Zoom Out."""
    def __init__(self, canvas, parent=None):
        super().__init__(canvas, parent)
        
        # Remove 'Configure subplots' action
        # It's usually the 7th action in standard toolbar, but let's search by text/tooltip
        actions = self.actions()
        for action in actions:
            if action.text() == 'Subplots' or 'subplot' in action.toolTip().lower():
                self.removeAction(action)
        
        # Add Zoom Out button
        self.zoom_out_act = QAction("Zoom Out", self)
        # Use a standard icon (Minimize button looks like a minus sign) as a proxy for Zoom Out
        self.zoom_out_act.setIcon(QApplication.style().standardIcon(QStyle.SP_TitleBarMinButton))
        self.zoom_out_act.setToolTip("Zoom Out (1.25x)")
        self.zoom_out_act.triggered.connect(self.zoom_out)
        
        # Update actions list after removal to ensure next_action is valid
        # (Using stale list causes insertAction to fail if next_action was the removed one)
        actions = self.actions()

        # Insert before the last separator (Save is usually last, or Configure subplots was near end)
        # Standard toolbar order: Home, Back, Forward, Pan, Zoom, Subplots, Save
        # We want Zoom Out next to Zoom.
        zoom_action = None
        for action in actions:
            if 'zoom' in action.toolTip().lower() and 'rect' in action.toolTip().lower():
                zoom_action = action
                break
        
        if zoom_action:
            # Insert after the existing Zoom button.
            found = False
            for i, action in enumerate(actions):
                if action == zoom_action:
                    # Insert before the next item in the UPDATED list
                    if i + 1 < len(actions):
                        next_action = actions[i+1]
                        self.insertAction(next_action, self.zoom_out_act)
                        found = True
                    break
            
            if not found:
                 self.addAction(self.zoom_out_act)
        else:
            self.addAction(self.zoom_out_act)

    
    def zoom_out(self):
        """Zoom out by expanding axes limits."""
        if not self.canvas.figure.axes:
            return
        ax = self.canvas.figure.axes[0]
        xlim = ax.get_xlim()
        ylim = ax.get_ylim()
        
        # Factor to expand > 1
        factor = 1.25
        x_center = (xlim[0] + xlim[1]) / 2
        y_center = (ylim[0] + ylim[1]) / 2
        
        new_w = (xlim[1] - xlim[0]) * factor
        new_h = (ylim[1] - ylim[0]) * factor
        
        ax.set_xlim([x_center - new_w/2, x_center + new_w/2])
        ax.set_ylim([y_center - new_h/2, y_center + new_h/2])
        
        self.canvas.draw()

class TestItemSpecDialog(QDialog):
    """Dialog to edit per-part per-test-item configuration: Spec Min/Max."""
    def __init__(self, parent, parts, test_items, specs):
        super().__init__(parent)
        self.setWindowTitle('Test Item Specification Setup')
        self.resize(1000, 400)
        self.parts = parts
        self.test_items = test_items
        self.specs = specs or {}
        layout = QVBoxLayout()
        
        # Rows: Parts, Cols: Test Items * 2 (Min, Max)
        col_count = 1 + len(test_items) * 2
        self.table = QTableWidget(len(parts), col_count)
        
        # Headers
        headers = ['Part Name']
        for item in test_items:
            headers.append(f"{item}\nMin")
            headers.append(f"{item}\nMax")
        self.table.setHorizontalHeaderLabels(headers)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        
        for r, part_name in enumerate(parts):
            # Part Name (Read Only)
            item = QTableWidgetItem(part_name)
            item.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)
            self.table.setItem(r, 0, item)
            
            part_specs = self.specs.get(part_name, {})
            
            for c, test_item in enumerate(test_items):
                # 0 is part name, so start at 1
                # test_item c maps to cols 1+2*c and 2+2*c
                col_min = 1 + c * 2
                col_max = 2 + c * 2
                
                item_specs = part_specs.get(test_item, {})
                mi = str(item_specs.get('min', ''))
                ma = str(item_specs.get('max', ''))
                
                self.table.setItem(r, col_min, QTableWidgetItem(mi))
                self.table.setItem(r, col_max, QTableWidgetItem(ma))
            
        layout.addWidget(self.table)
        
        btn_h = QHBoxLayout()
        self.save_btn = QPushButton('Save')
        self.save_btn.clicked.connect(self.on_save)
        btn_h.addWidget(self.save_btn)
        self.close_btn = QPushButton('Close')
        self.close_btn.clicked.connect(self.reject)
        btn_h.addWidget(self.close_btn)
        layout.addLayout(btn_h)
        self.setLayout(layout)

    def on_save(self):
        new_specs = {}
        for r in range(self.table.rowCount()):
            part_item = self.table.item(r, 0)
            if not part_item: continue
            part_name = part_item.text().strip()
            
            part_data = {}
            for c, test_item in enumerate(self.test_items):
                col_min = 1 + c * 2
                col_max = 2 + c * 2
                
                mi_item = self.table.item(r, col_min)
                ma_item = self.table.item(r, col_max)
                
                try:
                    mi_val = float(mi_item.text().strip()) if mi_item and mi_item.text().strip()!='' else 0.0
                except: mi_val = 0.0
                try:
                    ma_val = float(ma_item.text().strip()) if ma_item and ma_item.text().strip()!='' else 0.0
                except: ma_val = 0.0
                
                part_data[test_item] = {'min': mi_val, 'max': ma_val}
            
            new_specs[part_name] = part_data
            
        self.specs = new_specs
        
        # Save Test Item Specs
        try:
            cfg = get_config_file('test_item_specs_v2.json')
            with open(cfg, 'w', encoding='utf-8') as f:
                json.dump(self.specs, f, indent=2)
        except: pass
        
        QMessageBox.information(self.parent(), 'Saved', 'Saved Test Item Specifications')
        self.accept()

class TestItemRangeDialog(QDialog):
    """Dialog to edit per-part per-test-item configuration: Range Start/End (Time or Angle)."""
    def __init__(self, parent, parts, test_items, ranges, title='Test Item Range Setup', labels=('Start', 'End')):
        super().__init__(parent)
        self.setWindowTitle(title)
        self.resize(1100, 450)
        self.parts = parts
        self.test_items = test_items
        self.ranges = ranges or {}
        self.labels = labels
        layout = QVBoxLayout()
        
        # Rows: Parts, Cols: Test Items * 2 (Start, End)
        col_count = 1 + len(test_items) * 2
        self.table = QTableWidget(len(parts), col_count)
        
        # Headers
        headers = ['Part Name']
        for item in test_items:
            headers.append(f"{item}\n{labels[0]}")
            headers.append(f"{item}\n{labels[1]}")
        self.table.setHorizontalHeaderLabels(headers)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        
        for r, part_name in enumerate(parts):
            # Part Name (Read Only)
            item = QTableWidgetItem(part_name)
            item.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)
            self.table.setItem(r, 0, item)
            
            part_ranges = self.ranges.get(part_name, {})
            
            for c, test_item in enumerate(test_items):
                col_start = 1 + c * 2
                col_end = 2 + c * 2
                
                item_range = part_ranges.get(test_item, {})
                st = str(item_range.get('start', ''))
                en = str(item_range.get('end', ''))
                
                self.table.setItem(r, col_start, QTableWidgetItem(st))
                self.table.setItem(r, col_end, QTableWidgetItem(en))
            
        layout.addWidget(self.table)
        
        btn_h = QHBoxLayout()
        self.save_btn = QPushButton('Save')
        self.save_btn.clicked.connect(self.on_save)
        btn_h.addWidget(self.save_btn)
        self.close_btn = QPushButton('Close')
        self.close_btn.clicked.connect(self.reject)
        btn_h.addWidget(self.close_btn)
        layout.addLayout(btn_h)
        self.setLayout(layout)

    def on_save(self):
        new_ranges = {}
        for r in range(self.table.rowCount()):
            part_item = self.table.item(r, 0)
            if not part_item: continue
            part_name = part_item.text().strip()
            
            part_data = {}
            for c, test_item in enumerate(self.test_items):
                col_start = 1 + c * 2
                col_end = 2 + c * 2
                
                st_item = self.table.item(r, col_start)
                en_item = self.table.item(r, col_end)
                
                try:
                    st_val = float(st_item.text().strip()) if st_item and st_item.text().strip()!='' else 0.0
                except: st_val = 0.0
                try:
                    en_val = float(en_item.text().strip()) if en_item and en_item.text().strip()!='' else 0.0
                except: en_val = 0.0
                
                part_data[test_item] = {'start': st_val, 'end': en_val}
            
            new_ranges[part_name] = part_data
            
        self.ranges = new_ranges
        self.accept()

class PartConfigDialog(QDialog):
    """Dialog to edit per-part configuration: start/end ranges (Time or Angle)."""
    def __init__(self, parent, parts, ranges, title='Part Time Range Configuration', labels=('Range Start (s)', 'Range End (s)')):
        super().__init__(parent)
        self.setWindowTitle(title)
        self.resize(500, 350)
        self.parts = parts
        self.ranges = ranges or {}
        layout = QVBoxLayout()
        # Columns: Part Name, Start(s), End(s)
        self.table = QTableWidget(len(parts), 3)
        self.table.setHorizontalHeaderLabels(['Part Name', labels[0], labels[1]])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        for i, p in enumerate(parts):
            self.table.setItem(i, 0, QTableWidgetItem(p))
            # Make part name read-only
            self.table.item(i, 0).setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)

            # Range
            r = self.ranges.get(p, {})
            st = str(r.get('start', ''))
            en = str(r.get('end', ''))
            self.table.setItem(i, 1, QTableWidgetItem(st))
            self.table.setItem(i, 2, QTableWidgetItem(en))
            
        layout.addWidget(self.table)
        btn_h = QHBoxLayout()
        self.save_btn = QPushButton('Save')
        self.save_btn.clicked.connect(self.on_save)
        btn_h.addWidget(self.save_btn)
        self.close_btn = QPushButton('Close')
        self.close_btn.clicked.connect(self.reject)
        btn_h.addWidget(self.close_btn)
        layout.addLayout(btn_h)
        self.setLayout(layout)

    def on_save(self):
        new_ranges = {}
        for r in range(self.table.rowCount()):
            p_item = self.table.item(r, 0)
            if not p_item: continue
            p = p_item.text().strip()
            
            # 1. Range
            s_item = self.table.item(r, 1)
            e_item = self.table.item(r, 2)
            try:
                st = float(s_item.text().strip()) if s_item and s_item.text().strip()!='' else 0.0
            except: st = 0.0
            try:
                en = float(e_item.text().strip()) if e_item and e_item.text().strip()!='' else 0.0
            except: en = 0.0
            new_ranges[p] = {'start': st, 'end': en}
            
        self.ranges = new_ranges
        # Parent handles saving to specific file
        # QMessageBox.information(self.parent(), 'Saved', 'Saved Ranges')
        self.accept()

class CalibrationDialog(QDialog):
    """Dialog to edit per-part K factor configuration."""
    def __init__(self, parent, parts, calibration_data):
        super().__init__(parent)
        self.setWindowTitle('Calibration Setup')
        self.resize(450, 400)
        self.parts = parts
        self.calibration_data = calibration_data or {}
        layout = QVBoxLayout()
        # Columns: Part Name, Calibration
        self.table = QTableWidget(len(parts), 2)
        self.table.setHorizontalHeaderLabels(['Part Name', 'Calibration'])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        
        for r, part_name in enumerate(parts):
            # Part Name (Read Only)
            item = QTableWidgetItem(part_name)
            item.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)
            self.table.setItem(r, 0, item)
            
            # Factor K
            # Use stripped part name for key lookup to avoid mismatch
            key = part_name.strip()
            k_val = self.calibration_data.get(key, 1.0)
            spin = CommaDoubleSpinBox()
            spin.setRange(0.000001, 10000.0)
            spin.setDecimals(6)
            spin.setSingleStep(0.01)
            spin.setValue(float(k_val))
            self.table.setCellWidget(r, 1, spin)
            
        layout.addWidget(self.table)
        
        btn_h = QHBoxLayout()
        self.save_btn = QPushButton('Save')
        self.save_btn.clicked.connect(self.on_save)
        btn_h.addWidget(self.save_btn)
        self.close_btn = QPushButton('Close')
        self.close_btn.clicked.connect(self.reject)
        btn_h.addWidget(self.close_btn)
        layout.addLayout(btn_h)
        self.setLayout(layout)

    def on_save(self):
        new_data = {}
        for r in range(self.table.rowCount()):
            part_item = self.table.item(r, 0)
            if not part_item: continue
            part_name = part_item.text().strip()
            
            spin = self.table.cellWidget(r, 1)
            val = spin.value() if spin else 1.0
            new_data[part_name] = val
            
        self.calibration_data.update(new_data)
        
        # Save to calibration.json
        try:
            cfg = get_config_file('calibration.json')
            with open(cfg, 'w', encoding='utf-8') as f:
                json.dump(self.calibration_data, f, indent=2)
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to save calibration: {e}")
            return
        
        QMessageBox.information(self.parent(), 'Saved', 'Saved Calibration Data')
        self.accept()

class TorquePlotViewer(QMainWindow):
    """CSV Torque Plot Viewer"""
    
    def __init__(self):
        super().__init__()
        self.time_data = []
        self.torque_data = []
        # Support multiple datasets: list of dicts {name, time, torque, color}
        self.datasets = []
        # Color cycle from matplotlib
        self._colors = list(plt.rcParams.get('axes.prop_cycle').by_key().get('color', ['b', 'g', 'r', 'c', 'm', 'y', 'k']))
        self.file_path = ""
        # support multiple samples for export
        self.samples = []  # list of dicts: {'path','name','time','torque','part_no'}
        # Keep references to legend item widgets (dicts) so we can remove/update/reorder them
        self.files_legend_items = []
        # Consolidated Ranges: {Part: {TestItem: {start, end}}}
        self.test_item_time_ranges = {}
        try:
            cfg = get_config_file('test_item_time_ranges.json')
            if cfg.exists():
                with open(cfg, 'r', encoding='utf-8') as f:
                    self.test_item_time_ranges = json.load(f)
        except Exception: pass

        self.test_item_angle_ranges = {}
        try:
            cfg = get_config_file('test_item_angle_ranges.json')
            if cfg.exists():
                with open(cfg, 'r', encoding='utf-8') as f:
                    self.test_item_angle_ranges = json.load(f)
        except Exception: pass
        
        self.part_specs = {}
        try:
            cfg = get_config_file('part_specs.json')
            if cfg.exists():
                try:
                    with open(cfg, 'r', encoding='utf-8') as f:
                        self.part_specs = json.load(f)
                except: pass
        except Exception:
            self.part_specs = {}
        # Load Test Item Specs (v2 preferred)
        self.test_item_specs = {}
        try:
            cfg = get_config_file('test_item_specs_v2.json')
            if cfg.exists():
                try:
                    with open(cfg, 'r', encoding='utf-8') as f:
                        self.test_item_specs = json.load(f)
                except: pass
            else:
                # migration attempt or fallback? For now start empty if v2 missing, 
                # or maybe support old format? User wants new structure.
                # Let's clean start or check old file?
                # The old structure was {Item: {min,max}}, new is {Part: {Item: {min,max}}}.
                pass
        except Exception:
            self.test_item_specs = {}

        self.calibration_data = {}
        try:
            cfg = get_config_file('calibration.json')
            if cfg.exists():
                try:
                    with open(cfg, 'r', encoding='utf-8') as f:
                        self.calibration_data = json.load(f)
                except: pass
        except Exception:
            self.calibration_data = {}
        self.k_factor = 1.0

        self.csv_dir = ""
        self.report_dir = ""
        try:
            cfg = get_config_file('report_paths.json')
            if cfg.exists():
                with open(cfg, 'r', encoding='utf-8') as f:
                    pdata = json.load(f)
                    self.csv_dir = pdata.get('csv_dir', '')
                    self.report_dir = pdata.get('report_dir', '')
        except Exception:
            pass

        self.init_ui()
        # Ensure correct UI labels for default plot mode
        try:
            self.on_plot_mode_changed()
        except: pass

        # Initialize K factor and specs for the default selected part
        try:
             self.on_part_name_changed()
        except: pass

    def init_ui(self):
        """Initialize UI"""
        self.setWindowTitle("CSV Torque Plot Viewer")
        self.setGeometry(100, 100, 1450, 950)
        
        # Plot Viewer as central widget
        self.plot_tab = QWidget()
        self.setCentralWidget(self.plot_tab)
        
        central = self.plot_tab
        main_layout = QVBoxLayout(central)
        main_layout.setContentsMargins(6, 6, 6, 6)
        main_layout.setSpacing(6)
        
        # === Control Panel ===
        ctrl_group = QGroupBox("📁 File & Range")
        ctrl_group.setStyleSheet("QGroupBox{font-weight:bold;}")
        ctrl_group.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
        # Use a vertical layout with two rows and tighter spacing
        ctrl_v = QVBoxLayout()
        ctrl_v.setSpacing(6)
        ctrl_v.setContentsMargins(6, 6, 6, 6)

        # Top row: import buttons, file label, selector, export
        top_h = QHBoxLayout()
        self.import_btn = QPushButton("📂 Import CSV(s)")
        self.import_btn.setFixedHeight(28)
        self.import_btn.setStyleSheet("background-color: #4CAF50; color: white; font-weight: bold; padding: 4px;")
        self.import_btn.clicked.connect(self.import_csv)
        top_h.addWidget(self.import_btn)

        self.add_files_btn = QPushButton("➕ Add Files")
        self.add_files_btn.setFixedHeight(28)
        self.add_files_btn.setStyleSheet("background-color: #FF9800; color: white; font-weight: bold; padding: 4px;")
        self.add_files_btn.clicked.connect(self.add_files)
        top_h.addWidget(self.add_files_btn)

        self.clear_btn = QPushButton("🗑 Clear All")
        self.clear_btn.setFixedHeight(28)
        self.clear_btn.setStyleSheet("background-color: #F44336; color: white; font-weight: bold; padding: 4px;")
        self.clear_btn.clicked.connect(self.clear_all_samples)
        top_h.addWidget(self.clear_btn)

        # File label (shows last loaded file) - elide long paths
        self.file_label = QLabel("No file loaded")
        self.file_label.setStyleSheet("color: gray;")
        self.file_label.setMinimumWidth(200)
        self.file_label.setMaximumWidth(420)
        self.file_label.setWordWrap(False)
        self.file_label.setWordWrap(False)
        top_h.addWidget(self.file_label)
        top_h.addStretch(1)

        # Plot Mode
        # Group Mode label and combo tightly
        mode_h = QHBoxLayout()
        mode_h.setSpacing(5)
        mode_h.addWidget(QLabel("Mode:"))
        self.plot_mode_combo = QComboBox()
        self.plot_mode_combo.addItems(["Time vs Torque", "Angle vs Torque"])
        self.plot_mode_combo.currentIndexChanged.connect(self.on_plot_mode_changed)
        # Default index set at end of init_ui to ensure widgets exist
        mode_h.addWidget(self.plot_mode_combo)
        
        top_h.addLayout(mode_h)

        # Export button on the top right
        self.export_btn = QPushButton("📄 Export XLSX")
        self.export_btn.setFixedHeight(28)
        self.export_btn.setStyleSheet("background-color: #2196F3; color: white; font-weight: bold; padding: 4px;")
        self.export_btn.clicked.connect(self.export_xlsx)
        top_h.addWidget(self.export_btn)

        ctrl_v.addLayout(top_h)

        # Bottom row: STT range + Avg display
        bottom_h = QHBoxLayout()
        bottom_h.setSpacing(8)
        # STT Start/End removed from UI per user request; keep spinboxes
        # available internally but do not add their widgets to the layout.
        self.start_spin = QSpinBox()
        self.start_spin.setRange(1, 999999)
        self.start_spin.setValue(1)
        self.start_spin.setMaximumWidth(90)
        self.start_spin.valueChanged.connect(self.update_average)
        # (start_spin not added to layout)
        self.end_spin = QSpinBox()
        self.end_spin.setRange(1, 999999)
        self.end_spin.setValue(100)
        self.end_spin.setMaximumWidth(90)
        self.end_spin.valueChanged.connect(self.update_average)
        # (end_spin not added to layout)

        # Data range mode: Default / Manual (compact)
        bottom_h.addWidget(QLabel("Range:"))
        self.range_mode_combo = QComboBox()
        self.range_mode_combo.addItems(["Default", "Manual"])
        self.range_mode_combo.setMaximumWidth(90)
        self.range_mode_combo.currentIndexChanged.connect(self.on_range_mode_changed)
        bottom_h.addWidget(self.range_mode_combo)

        # Manual time inputs (hidden by default)
        self.start_time_spin = None
        self.end_time_spin = None
        try:
            from PyQt5.QtWidgets import QDoubleSpinBox
            self.start_time_spin = QDoubleSpinBox()
            self.start_time_spin.setDecimals(6)
            # Allow convenient fractional seconds entry (microsecond precision)
            try:
                self.start_time_spin.setSingleStep(1e-6)
            except Exception:
                try:
                    self.start_time_spin.setSingleStep(0.000001)
                except Exception:
                    pass
            self.start_time_spin.setRange(-1e9, 1e9)
            self.start_time_spin.setMaximumWidth(100)
            self.start_time_spin.setSuffix(" s")
            self.start_time_spin.setValue(0.0)
            self.start_time_spin.valueChanged.connect(self.update_average)
            self.end_time_spin = QDoubleSpinBox()
            self.end_time_spin.setDecimals(6)
            try:
                self.end_time_spin.setSingleStep(1e-6)
            except Exception:
                try:
                    self.end_time_spin.setSingleStep(0.000001)
                except Exception:
                    pass

            # Normalize input so both comma and dot decimal separators are accepted.
            try:
                def _normalize_spin_input(spin):
                    try:
                        txt = spin.lineEdit().text()
                        if not txt:
                            return
                        # Accept both comma and dot as decimal separators
                        txt2 = txt.replace(',', '.')
                        try:
                            v = float(txt2)
                            # Set numeric value
                            spin.setValue(v)
                            # Format display: up to 6 decimals, strip trailing zeros
                            disp = f"{v:.6f}".rstrip('0').rstrip('.')
                            # Use comma as decimal separator in the UI per user preference
                            disp = disp.replace('.', ',')
                            try:
                                # ensure the underlying QLineEdit shows the formatted text
                                spin.lineEdit().setText(disp)
                            except Exception:
                                pass
                        except Exception:
                            pass
                    except Exception:
                        pass

                try:
                    self.start_time_spin.editingFinished.connect(lambda sp=self.start_time_spin: _normalize_spin_input(sp))
                except Exception:
                    pass
                try:
                    self.end_time_spin.editingFinished.connect(lambda sp=self.end_time_spin: _normalize_spin_input(sp))
                except Exception:
                    pass

                # Also convert '.' to ',' while typing so the underlying
                # QDoubleSpinBox (which may use locale with comma decimal)
                # interprets input correctly. Preserve cursor position.
                try:
                    def _on_text_changed(spin, text):
                        try:
                            if '.' in text:
                                le = spin.lineEdit()
                                pos = le.cursorPosition()
                                new = text.replace('.', ',')
                                le.blockSignals(True)
                                le.setText(new)
                                # adjust cursor if needed
                                try:
                                    le.setCursorPosition(min(pos, len(new)))
                                except Exception:
                                    pass
                                le.blockSignals(False)
                        except Exception:
                            pass

                    try:
                        self.start_time_spin.lineEdit().textChanged.connect(lambda txt, sp=self.start_time_spin: _on_text_changed(sp, txt))
                    except Exception:
                        pass
                    try:
                        self.end_time_spin.lineEdit().textChanged.connect(lambda txt, sp=self.end_time_spin: _on_text_changed(sp, txt))
                    except Exception:
                        pass
                except Exception:
                    pass
            except Exception:
                pass
            self.end_time_spin.setRange(-1e9, 1e9)
            self.end_time_spin.setMaximumWidth(100)
            self.end_time_spin.setSuffix(" s")
            self.end_time_spin.setValue(1.0)
            self.end_time_spin.valueChanged.connect(self.update_average)
            # hide initially
            self.start_label_widget = QLabel("Start Time (s):")
            self.end_label_widget = QLabel("End Time (s):")
            self.start_label_widget.hide()
            self.end_label_widget.hide()
            self.start_time_spin.hide()
            self.end_time_spin.hide()
            
            bottom_h.addWidget(self.start_label_widget)
            bottom_h.addWidget(self.start_time_spin)
            bottom_h.addWidget(self.end_label_widget)
            bottom_h.addWidget(self.end_time_spin)
        except Exception:
            self.start_time_spin = None
            self.end_time_spin = None

        # Setup buttons for part time ranges (compact)
        self.range_setup_btn = QPushButton("Ranges")
        self.range_setup_btn.setMaximumWidth(80)
        self.range_setup_btn.setToolTip("Configure Test Item Ranges (Time/Angle)")
        self.range_setup_btn.clicked.connect(self.open_part_range_setup)
        bottom_h.addWidget(self.range_setup_btn)

        self.spec_setup_btn = QPushButton("Specs")
        self.spec_setup_btn.setMaximumWidth(70)
        self.spec_setup_btn.setToolTip("Configure Test Item Specifications")
        self.spec_setup_btn.clicked.connect(self.open_test_item_spec_setup)
        bottom_h.addWidget(self.spec_setup_btn)

        self.calibration_btn = QPushButton("Calibration")
        self.calibration_btn.setMaximumWidth(90)
        self.calibration_btn.setToolTip("Configure Calibration per Part")
        self.calibration_btn.clicked.connect(self.open_calibration_setup)
        bottom_h.addWidget(self.calibration_btn)

        bottom_h.addStretch()
        # Avg label moved to Info Panel to save space
        ctrl_v.addLayout(bottom_h)

        ctrl_group.setLayout(ctrl_v)
        main_layout.addWidget(ctrl_group)
        # === Files legend (color mapping) ===
        files_group = QGroupBox("📂 Imported Files")
        # files_group.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)

        files_layout = QVBoxLayout()
        files_layout.setContentsMargins(6, 6, 6, 6)
        
        self.files_legend_widget = QWidget()
        self.files_legend_layout = FlowLayout(self.files_legend_widget)
        self.files_legend_layout.setContentsMargins(2, 2, 2, 2)
        self.files_legend_layout.setSpacing(4)
        self.files_legend_widget.setLayout(self.files_legend_layout)
        
        # Use a scroll area for the file list so it doesn't push down Report Info
        scroll_files = QScrollArea()
        scroll_files.setWidgetResizable(True)
        scroll_files.setWidget(self.files_legend_widget)
        # Remove border from scroll area to blend in
        scroll_files.setFrameShape(QFrame.NoFrame)
        files_layout.addWidget(scroll_files)

        # Controls: single toggle select + compact remove/clear to minimize buttons
        files_ctrl_h = QHBoxLayout()
        files_ctrl_h.setSpacing(6)
        # compact: no global select/clear buttons; delete available per-item
        files_layout.addLayout(files_ctrl_h)
        files_group.setLayout(files_layout)
        # Restrict max height so it scrolls internally instead of growing
        files_group.setMaximumHeight(200)

        # NOTE: do not add files_group to main_layout here — it will be placed
        # into the left column alongside the metadata so it only occupies the
        # left side and leaves the plot area wide.
        
        # === Report Metadata ===
        meta_group = QGroupBox("📝 Report Info")
        meta_layout = QGridLayout()

        meta_layout.addWidget(QLabel("TEST ITEM:"), 0, 0)
        self.test_item_combo = QComboBox()
        self.test_item_combo.addItems(["Breakaway Torque", "Operating Torque", "Oscillating Torque"])
        self.test_item_combo.setEnabled(False) # Locked editing, synced from Thu thap
        try:
            self.test_item_combo.currentIndexChanged.connect(self.on_test_item_changed)
        except: pass
        meta_layout.addWidget(self.test_item_combo, 0, 1)

        meta_layout.addWidget(QLabel("PART NAME:"), 0, 2)
        self.part_name_combo = QComboBox()
        self.part_name_combo.addItems(["Inner Tie Rod", "Ball Joint", "Outer Tie Rod", "Stabilizer Link"])
        self.part_name_combo.setEnabled(False) # Locked editing, synced from Thu thap
        meta_layout.addWidget(self.part_name_combo, 0, 3)
        try:
            self.part_name_combo.currentIndexChanged.connect(self.on_part_name_changed)
        except Exception:
            pass

        meta_layout.addWidget(QLabel("PART NO:"), 1, 0)
        self.part_no_edit = QLineEdit("")
        self.part_no_edit.setMaximumWidth(140)
        def _on_part_no_changed(text):
            pos = self.part_no_edit.cursorPosition()
            upp = text.upper()
            if text != upp:
                self.part_no_edit.blockSignals(True)
                self.part_no_edit.setText(upp)
                self.part_no_edit.setCursorPosition(pos)
                self.part_no_edit.blockSignals(False)
        self.part_no_edit.textChanged.connect(_on_part_no_changed)
        part_sample_h = QHBoxLayout()
        part_sample_h.setSpacing(8)
        part_sample_h.addWidget(self.part_no_edit)
        part_sample_h.addWidget(QLabel("SAMPLE NO:"))
        self.sample_no_spin = QSpinBox()
        self.sample_no_spin.setRange(1, 99)
        self.sample_no_spin.setValue(1)
        self.sample_no_spin.setMaximumWidth(70)
        self.sample_no_spin.setToolTip("Sample No, valid range 01 to 99")
        part_sample_h.addWidget(self.sample_no_spin)
        meta_layout.addLayout(part_sample_h, 1, 1)

        # Report role fields (Write / Review / Approval) placed in a single horizontal row (compact)
        report_h = QHBoxLayout()
        report_h.setSpacing(8)
        report_h.addWidget(QLabel("Write:"))
        self.write_edit = QLineEdit("")
        self.write_edit.setMaximumWidth(140)
        report_h.addWidget(self.write_edit)
        report_h.addWidget(QLabel("Review:"))
        self.review_edit = QLineEdit("")
        self.review_edit.setMaximumWidth(140)
        report_h.addWidget(self.review_edit)
        report_h.addWidget(QLabel("Approval:"))
        self.approval_edit = QLineEdit("")
        self.approval_edit.setMaximumWidth(140)
        report_h.addWidget(self.approval_edit)
        meta_layout.addLayout(report_h, 3, 0, 1, 4)

        # SPECIFICATION moved to its own row below PART NO
        meta_layout.addWidget(QLabel("SPECIFICATION:"), 2, 0)
        spec_h = QHBoxLayout()
        self.spec_min_spin = None
        self.spec_max_spin = None
        try:
            self.spec_min_spin = CommaDoubleSpinBox()
            self.spec_min_spin.setRange(-1e6, 1e6)
            self.spec_min_spin.setMaximumWidth(120)
            self.spec_min_spin.setSuffix(" Nm")
            self.spec_min_spin.setValue(0.0)
            self.spec_min_spin.setToolTip('Specification minimum (Nm)')
            spec_h.addWidget(QLabel("Min:"))
            spec_h.addWidget(self.spec_min_spin)
            try:
                self.spec_min_spin.valueChanged.connect(self.update_average)
                self.spec_min_spin.editingFinished.connect(self.save_current_spec)
            except Exception:
                pass
            self.spec_max_spin = CommaDoubleSpinBox()
            self.spec_max_spin.setRange(-1e6, 1e6)
            self.spec_max_spin.setMaximumWidth(120)
            self.spec_max_spin.setSuffix(" Nm")
            self.spec_max_spin.setValue(0.0)
            self.spec_max_spin.setToolTip('Specification maximum (Nm)')
            spec_h.addWidget(QLabel("Max:"))
            spec_h.addWidget(self.spec_max_spin)
            try:
                self.spec_max_spin.valueChanged.connect(self.update_average)
                self.spec_max_spin.editingFinished.connect(self.save_current_spec)
            except Exception:
                pass
            self.spec_min_spin.show()
            self.spec_max_spin.show()
        except Exception:
            pass
            self.spec_edit = QLineEdit("")
            self.spec_edit.setMaximumWidth(180)
            spec_h.addWidget(self.spec_edit)
            try:
                self.spec_edit.textChanged.connect(self.update_average)
            except Exception:
                pass
        meta_layout.addLayout(spec_h, 2, 1, 1, 3)

        meta_layout.addWidget(QLabel("DATE:"), 4, 0)
        self.date_edit = QDateEdit(QDate.currentDate())
        self.date_edit.setCalendarPopup(True)
        self.date_edit.setMaximumWidth(140)
        meta_layout.addWidget(self.date_edit, 4, 1)

        meta_layout.addWidget(QLabel("TESTER:"), 4, 2)
        self.tester_edit = QLineEdit("")
        self.tester_edit.setMaximumWidth(180)
        def _on_tester_changed(text):
            pos = self.tester_edit.cursorPosition()
            clean = remove_diacritics_no_strip(text)
            if text != clean:
                self.tester_edit.blockSignals(True)
                self.tester_edit.setText(clean)
                self.tester_edit.setCursorPosition(pos)
                self.tester_edit.blockSignals(False)
        self.tester_edit.textChanged.connect(_on_tester_changed)
        meta_layout.addWidget(self.tester_edit, 4, 3)

        meta_layout.addWidget(QLabel("TEST PURPOSE:"), 5, 0)
        self.test_purpose_combo = QComboBox()
        self.test_purpose_combo.addItems([
            "Setting (S)",
            "First (F)",
            "Middle (M)",
            "Final (Z)",
            "Development (D)",
            "Long-term (L)",
            "other (O)"
        ])
        self.test_purpose_combo.setMaximumWidth(240)
        self.test_purpose_combo.currentIndexChanged.connect(self.on_test_purpose_changed)
        
        self.test_purpose_other_edit = QLineEdit("")
        self.test_purpose_other_edit.setPlaceholderText("Enter test purpose...")
        self.test_purpose_other_edit.hide()
        
        tp_container = QWidget()
        tp_layout = QHBoxLayout(tp_container)
        tp_layout.setContentsMargins(0,0,0,0)
        tp_layout.addWidget(self.test_purpose_combo)
        tp_layout.addWidget(self.test_purpose_other_edit)
        meta_layout.addWidget(tp_container, 5, 1)
        
        # Judgment display label (show OK/NG)
        meta_layout.addWidget(QLabel("JUDGMENT:"), 5, 2)
        self.judgment_h = QHBoxLayout()
        self.judgment_label = QLabel("")
        self.judgment_label.setFont(QFont("Arial", 11, QFont.Bold))
        self.judgment_label.setMaximumWidth(100)
        self.judgment_h.addWidget(self.judgment_label)
        meta_layout.addLayout(self.judgment_h, 5, 3)
        
        # Aspect ratio and Quantity together on Row 6 Col 0-1
        meta_layout.addWidget(QLabel("Graph H/W Ratio:"), 6, 0)
        ratio_qty_h = QHBoxLayout()
        ratio_qty_h.setSpacing(4)
        ratio_qty_h.setContentsMargins(0, 0, 0, 0)
        
        self.aspect_ratio_spin = CommaDoubleSpinBox()
        self.aspect_ratio_spin.setRange(0.1, 2.0)
        self.aspect_ratio_spin.setDecimals(2)
        self.aspect_ratio_spin.setSingleStep(0.05)
        self.aspect_ratio_spin.setValue(0.75)
        self.aspect_ratio_spin.setToolTip("Height to Width ratio for exported graph (0.75 = 3/4)")
        self.aspect_ratio_spin.setMaximumWidth(90)
        ratio_qty_h.addWidget(self.aspect_ratio_spin)
        
        ratio_qty_h.addWidget(QLabel("Qty:"))
        self.quantity_spin = QSpinBox()
        self.quantity_spin.setRange(1, 1000000)
        self.quantity_spin.setValue(1)
        self.quantity_spin.setMaximumWidth(70)
        ratio_qty_h.addWidget(self.quantity_spin)
        
        ratio_qty_widget = QWidget()
        ratio_qty_widget.setLayout(ratio_qty_h)
        meta_layout.addWidget(ratio_qty_widget, 6, 1)

        # Team combo on Row 6 Col 2-3
        meta_layout.addWidget(QLabel("TEAM:"), 6, 2)
        self.team_combo = QComboBox()
        self.team_combo.addItems(TEAMS)
        self.team_combo.setMaximumWidth(180)
        meta_layout.addWidget(self.team_combo, 6, 3)

        # Remark on Row 7 Col 0-3. Not persisted; cleared when returning to acquisition tab.
        meta_layout.addWidget(QLabel("REMARK:"), 7, 0)
        self.remark_edit = QLineEdit("")
        self.remark_edit.setMaxLength(100)
        self.remark_edit.setPlaceholderText("MAX 100 CHARACTERS")
        def _on_remark_changed(text):
            pos = self.remark_edit.cursorPosition()
            upp = text.upper()
            if text != upp:
                self.remark_edit.blockSignals(True)
                self.remark_edit.setText(upp)
                self.remark_edit.setCursorPosition(min(pos, len(upp)))
                self.remark_edit.blockSignals(False)
        self.remark_edit.textChanged.connect(_on_remark_changed)
        meta_layout.addWidget(self.remark_edit, 7, 1, 1, 3)

        # Lot No on Row 8 Col 0-1
        meta_layout.addWidget(QLabel("LOT NO:"), 8, 0)
        self.lot_no_edit = QLineEdit("")
        self.lot_no_edit.setMaximumWidth(180)
        meta_layout.addWidget(self.lot_no_edit, 8, 1)

        # Line No combo on Row 8 Col 2-3
        meta_layout.addWidget(QLabel("LINE NO:"), 8, 2)
        self.line_no_combo = QComboBox()
        self.line_no_combo.addItems(LINE_NOS)
        self.line_no_combo.setMaximumWidth(180)
        meta_layout.addWidget(self.line_no_combo, 8, 3)

        # Profile save/load buttons (Row 9)
        prof_h = QHBoxLayout()
        prof_h.setSpacing(8)
        self.save_profile_btn = QPushButton("Save Profile")
        self.save_profile_btn.setFixedHeight(26)
        self.save_profile_btn.clicked.connect(self.save_profile)
        prof_h.addWidget(self.save_profile_btn)
        self.load_profile_btn = QPushButton("Load Profile")
        self.load_profile_btn.setFixedHeight(26)
        self.load_profile_btn.clicked.connect(self.load_profile)
        prof_h.addWidget(self.load_profile_btn)
        meta_layout.addLayout(prof_h, 9, 0, 1, 4)

        # Folder Paths (Row 10 & 11)
        meta_layout.addWidget(QLabel("CSV File path:"), 10, 0)
        csv_path_h = QHBoxLayout()
        csv_path_h.setSpacing(4)
        csv_path_h.setContentsMargins(0, 0, 0, 0)
        self.csv_path_edit = QLineEdit(self.csv_dir)
        csv_path_h.addWidget(self.csv_path_edit)
        self.csv_browse_btn = QPushButton()
        self.csv_browse_btn.setIcon(self.style().standardIcon(QStyle.SP_DirOpenIcon))
        self.csv_browse_btn.setToolTip("Browse CSV save directory")
        self.csv_browse_btn.setFixedSize(34, 30)
        self.csv_browse_btn.setIconSize(self.csv_browse_btn.size() * 0.62)
        self.csv_browse_btn.clicked.connect(self.browse_csv_path)
        csv_path_h.addWidget(self.csv_browse_btn)
        csv_path_widget = QWidget()
        csv_path_widget.setLayout(csv_path_h)
        meta_layout.addWidget(csv_path_widget, 10, 1, 1, 3)

        meta_layout.addWidget(QLabel("Report File path:"), 11, 0)
        report_path_h = QHBoxLayout()
        report_path_h.setSpacing(4)
        report_path_h.setContentsMargins(0, 0, 0, 0)
        self.report_path_edit = QLineEdit(self.report_dir)
        report_path_h.addWidget(self.report_path_edit)
        self.report_browse_btn = QPushButton()
        self.report_browse_btn.setIcon(self.style().standardIcon(QStyle.SP_DirOpenIcon))
        self.report_browse_btn.setToolTip("Browse report save directory")
        self.report_browse_btn.setFixedSize(34, 30)
        self.report_browse_btn.setIconSize(self.report_browse_btn.size() * 0.62)
        self.report_browse_btn.clicked.connect(self.browse_report_path)
        report_path_h.addWidget(self.report_browse_btn)
        report_path_widget = QWidget()
        report_path_widget.setLayout(report_path_h)
        meta_layout.addWidget(report_path_widget, 11, 1, 1, 3)

        # Save Report Button (Row 12)
        self.save_report_btn = QPushButton("💾 Save the Report")
        self.save_report_btn.setFixedHeight(32)
        self.save_report_btn.setStyleSheet("background-color: #4CAF50; color: white; font-weight: bold; font-size: 10pt;")
        self.save_report_btn.clicked.connect(self.save_report)
        meta_layout.addWidget(self.save_report_btn, 12, 0, 1, 4)

        # Place judgment in its own column so it's always visible
        meta_layout.addWidget(QLabel("JUDGMENT:"), 5, 2)
        meta_layout.addWidget(self.judgment_label, 5, 3)
        meta_group.setLayout(meta_layout)
        # meta_group.setMaximumWidth(380) => Removed

        # === Plot ===
        self.plot_group = QGroupBox("📈 Plot")  # Generic title, will be updated by on_plot_mode_changed
        plot_layout = QVBoxLayout()
        plot_layout.setContentsMargins(6, 6, 6, 6)
        plot_layout.setSpacing(6)
        
        self.fig = Figure(figsize=(9, 5), dpi=100)
        self.ax = self.fig.add_subplot(111)
        self.canvas = FigureCanvas(self.fig)
        self.canvas.setMinimumHeight(420)
        
        # Initial labels will be set by on_plot_mode_changed
        self.ax.set_ylabel("Torque (N·m)", fontsize=10)
        self.ax.grid(True, alpha=0.3)
        self.fig.tight_layout()
        
        # Add matplotlib navigation toolbar for zoom/pan
        # Add matplotlib navigation toolbar for zoom/pan
        self.toolbar = CustomToolbar(self.canvas, self)

        self.toolbar.setMaximumHeight(32)
        plot_layout.addWidget(self.toolbar)
        plot_layout.addWidget(self.canvas)
        self.plot_group.setLayout(plot_layout)

        # === Layout Assembly with QSplitter for Responsiveness ===
        
        # Left Side Container (Scrollable)
        left_container = QWidget()
        left_layout = QVBoxLayout(left_container)
        left_layout.setContentsMargins(4, 4, 10, 4)
        left_layout.setSpacing(10)
        
        # Add Widgets to Left Side
        left_layout.addWidget(files_group)
        
        # Report title
        title_h = QHBoxLayout()
        title_h.setSpacing(6)
        title_h.addWidget(QLabel("Report Title:"))
        self.report_title_edit = QLineEdit("TEST REPORT")
        title_h.addWidget(self.report_title_edit)
        left_layout.addLayout(title_h)
        
        # Metadata
        # Remove fixed width constraints to allow fit
        meta_group.setMaximumWidth(16777215) 
        meta_group.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
        left_layout.addWidget(meta_group)

        # === Cycle Selection ===
        self.cycle_group = QGroupBox("🔄 Cycle Selection")
        self.cycle_group.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
        cycle_v = QVBoxLayout()

        # 1. Target File Selection (Moved from top to here for better UX)
        # 1. Target File Selection (Moved from top to here for better UX)
        file_sel_h = QHBoxLayout()
        file_sel_h.setSpacing(5) # Tighter spacing
        file_sel_h.addWidget(QLabel("File:"))
        self.file_select_combo = QComboBox()
        self.file_select_combo.addItem("All")
        self.file_select_combo.currentIndexChanged.connect(self.on_file_selection_changed)
        file_sel_h.addWidget(self.file_select_combo)
        # Add stretch to keep them close to the left/start if the combo isn't expanding, or just let combo expand?
        # User said "hơi xa nhau", implies too much space between label and combo?
        # Or combo is too far right? 
        # Typically QComboBox expands. If we want label and combo close, we usually just reduce spacing.
        # But if the layout justifies them apart, addStretch at end.
        # Let's add stretch at the end to force them left-aligned and close.
        # However, usually we want the combo to fill available width. 
        # If user says "xa nhau" perhaps the label is left aligned and combo is right aligned? 
        # Let's assume standard left-to-right flow. Reducing spacing helps.
        # If the combo is expanding to fill a wide groupbox, the gap might appear large if layout is "Justify" or similar (unlikely default).
        # We'll just set spacing(5).
        cycle_v.addLayout(file_sel_h)
        
        # 2. Cycle Controls
        cycle_btns_h = QHBoxLayout()
        self.btn_cycle_all = QPushButton("All")
        self.btn_cycle_all.setFixedWidth(70)
        self.btn_cycle_all.clicked.connect(self.select_all_cycles)
        self.btn_cycle_none = QPushButton("None")
        self.btn_cycle_none.setFixedWidth(70)
        self.btn_cycle_none.clicked.connect(self.deselect_all_cycles)
        cycle_btns_h.addWidget(self.btn_cycle_all)
        cycle_btns_h.addWidget(self.btn_cycle_none)
        cycle_btns_h.addStretch()
        cycle_v.addLayout(cycle_btns_h)
        
        self.cycle_layout = FlowLayout()
        self.cycle_layout.setContentsMargins(0, 5, 0, 5)
        cycle_v.addLayout(self.cycle_layout)
        self.cycle_group.setLayout(cycle_v)
        self.cycle_checkboxes = []
        left_layout.addWidget(self.cycle_group)

        # === Info Panel (Vertical Grid for Side Panel) ===
        info_group = QGroupBox("📊 Data Info")
        info_group.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
        info_group.setStyleSheet("QGroupBox { background-color: #fafafa; border: 1px solid #dcdcdc; border-radius: 4px; margin-top: 6px; padding-top: 4px; } QGroupBox::title { color: #555; font-size: 11px; font-weight: bold; }")
        
        # Use Grid layout for clearer vertical presentation in side panel
        info_layout = QGridLayout()
        info_layout.setContentsMargins(8, 8, 8, 8)
        info_layout.setSpacing(6)
        
        # Row 0: Source Label (All or specific file)
        self.source_label = QLabel("Target: All")
        self.source_label.setStyleSheet("color: #1976D2; font-weight: bold; font-size: 11px;")
        info_layout.addWidget(self.source_label, 0, 0, 1, 4) # Span across row

        # Row 1: Samples & Points
        self.samples_label = QLabel("Samples: 0")
        self.samples_label.setStyleSheet("color: #555; font-size: 11px;")
        info_layout.addWidget(self.samples_label, 1, 0)
        
        self.total_label = QLabel("Pts: 0")
        self.total_label.setStyleSheet("font-weight: bold; color: #333; font-size: 11px;")
        info_layout.addWidget(self.total_label, 1, 1)

        # Row 2: Avg
        info_layout.addWidget(QLabel("Avg (Nm):"), 2, 0)
        self.avg_label = QLabel("-")
        self.avg_label.setStyleSheet("color: #F44336; font-weight: bold; font-size: 14px;")
        info_layout.addWidget(self.avg_label, 2, 1)
        
        info_layout.addWidget(QLabel("Avg (Kgf.cm):"), 2, 2)
        self.avg_label_kgf = QLabel("-")
        self.avg_label_kgf.setStyleSheet("color: #F44336; font-weight: bold; font-size: 14px;")
        info_layout.addWidget(self.avg_label_kgf, 2, 3)
        
        # Row 3: Min
        info_layout.addWidget(QLabel("Min (Nm):"), 3, 0)
        self.min_label = QLabel("-")
        self.min_label.setStyleSheet("color: #D32F2F; font-size: 11px;")
        info_layout.addWidget(self.min_label, 3, 1)
        
        info_layout.addWidget(QLabel("Min (Kgf.cm):"), 3, 2)
        self.min_label_kgf = QLabel("-")
        self.min_label_kgf.setStyleSheet("color: #D32F2F; font-size: 11px;")
        info_layout.addWidget(self.min_label_kgf, 3, 3)
        
        # Row 4: Max
        info_layout.addWidget(QLabel("Max (Nm):"), 4, 0)
        self.max_label = QLabel("-")
        self.max_label.setStyleSheet("color: #1976D2; font-size: 11px;")
        info_layout.addWidget(self.max_label, 4, 1)
        
        info_layout.addWidget(QLabel("Max (Kgf.cm):"), 4, 2)
        self.max_label_kgf = QLabel("-")
        self.max_label_kgf.setStyleSheet("color: #1976D2; font-size: 11px;")
        info_layout.addWidget(self.max_label_kgf, 4, 3)

        info_group.setLayout(info_layout)
        # Add to left layout (instead of main_layout bottom)
        left_layout.addWidget(info_group)

        left_layout.addStretch() # Push everything up
        
        # Scroll Area for Left Side
        left_scroll = QScrollArea()
        left_scroll.setWidgetResizable(True)
        left_scroll.setWidget(left_container)
        left_scroll.setFrameShape(QFrame.NoFrame)
        left_scroll.setMinimumWidth(530)
        # Style the scroll area to match background
        left_scroll.setStyleSheet("QScrollArea { background-color: transparent; border: none; }")
        
        # Splitter
        self.splitter = QSplitter(Qt.Horizontal)
        self.splitter.addWidget(left_scroll)
        self.splitter.addWidget(self.plot_group)  # Use self.plot_group
        
        # Set initial sizes (Left=450, Right=Remaining)
        self.splitter.setSizes([450, 1000])


        self.splitter.setCollapsible(0, False) # Left side always visible (min width) but can be resized
        self.splitter.setCollapsible(1, False)
        
        main_layout.addWidget(self.splitter)
        
        # Data Info moved to Left Side


        # Apply Global Stylesheet
        self.setStyleSheet("""
            QMainWindow {
                background-color: #f4f6f9;
            }
            QGroupBox {
                background-color: #ffffff;
                border: 1px solid #e0e0e0;
                border-radius: 8px;
                margin-top: 1.2em;
                font-family: "Segoe UI", Arial, sans-serif;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                subcontrol-position: top left;
                left: 10px;
                padding: 0 5px;
                color: #2c3e50;
                font-weight: bold;
                font-size: 13px;
            }
            QPushButton {
                border-radius: 5px;
                padding: 6px 12px;
                font-family: "Segoe UI", sans-serif;
                font-weight: 600;
                border: 1px solid #dcdcdc;
                background-color: #ffffff;
            }
            QPushButton:hover {
                background-color: #f0f0f0;
                border: 1px solid #bbbbbb;
            }
            QPushButton:pressed {
                background-color: #e0e0e0;
            }
            QLabel {
                font-family: "Segoe UI", sans-serif;
                font-size: 12px;
                color: #333;
            }
            QLineEdit, QComboBox, QSpinBox, QDoubleSpinBox, QDateEdit {
                padding: 5px;
                border: 1px solid #cccccc;
                border-radius: 4px;
                background-color: #ffffff;
                selection-background-color: #3f51b5;
            }
            QLineEdit:focus, QComboBox:focus, QSpinBox:focus, QDoubleSpinBox:focus {
                border: 1px solid #3f51b5;
            }
            QSplitter::handle {
                background-color: #dcdcdc;
                width: 2px;
            }
            QScrollBar:vertical {
                border: none;
                background: #f0f0f0;
                width: 10px;
                margin: 0px 0px 0px 0px;
            }
            QScrollBar::handle:vertical {
                background: #c0c0c0;
                min-height: 20px;
                border-radius: 5px;
            }
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {
                height: 0px;
            }
        """)

        # ===== Application / Statusbar Icon =====
        try:
            # Support PyInstaller by checking _MEIPASS
            base_path = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
            icon_path = os.path.join(base_path, 'data-analysis.png')
            if os.path.exists(icon_path):
                pix = QPixmap(icon_path)
                if not pix.isNull():
                    # Set application and window icon (larger)
                    try:
                        app_icon = QIcon(pix)
                        QApplication.setWindowIcon(app_icon)
                    except Exception:
                        pass
                    try:
                        self.setWindowIcon(QIcon(pix))
                    except Exception:
                        pass
                    # Add a small icon to the status bar
                    try:
                        status_lbl = QLabel()
                        small = pix.scaled(16, 16, Qt.KeepAspectRatio, Qt.SmoothTransformation)
                        status_lbl.setPixmap(small)
                        status_lbl.setToolTip('Data Analysis')
                        # Ensure status bar exists and add as permanent widget
                        try:
                            sb = self.statusBar()
                        except Exception:
                            sb = None
                        if sb is None:
                            # Create a status bar and set it
                            try:
                                from PyQt5.QtWidgets import QStatusBar
                                sb = QStatusBar()
                                self.setStatusBar(sb)
                            except Exception:
                                sb = None
                        if sb is not None:
                            sb.addPermanentWidget(status_lbl)
                    except Exception:
                        pass
        except Exception:
            pass
        
        # Trigger initial mode update to sync UI (labels, plot config)
        # Set to Angle vs Torque by default
        self.plot_mode_combo.setCurrentIndex(1)
        self.on_plot_mode_changed()
    
    def import_csv(self):
        """Import one or more CSV files and add them as samples."""
        paths, _ = QFileDialog.getOpenFileNames(self, "Open CSV File(s)", "", "CSV Files (*.csv);;All Files (*)")
        if not paths:
            return

        added = 0
        for path in paths:
            try:
                tlist, trlist, alist, clist = self.read_csv_arrays(path)
                if not tlist:
                    continue
                
                # Check Mode Restrictions
                mode = self.plot_mode_combo.currentText()
                is_angle = (mode == "Angle vs Torque")
                if is_angle:
                     angle_range = max(alist) - min(alist) if alist else 0
                     if angle_range == 0.0 and all(a == 0.0 for a in alist):
                         # Skip this file if in Angle mode and no angle data
                         continue

                name = os.path.basename(path)
                color = self._colors[len(self.samples) % len(self._colors)]
                # PART NO must come from the UI; if empty, keep empty string (no fallback to filename)
                part_no_val = self.part_no_edit.text().strip()
                sample = {
                    'path': path, 
                    'name': name, 
                    'time': tlist, 
                    'torque': trlist, 
                    'angle': alist, 
                    'cycle': clist, 
                    'selected_cycles': sorted(list(set(clist))) if clist else [1],
                    'part_no': part_no_val
                }
                self.samples.append(sample)
                # add to selector and legend; legend shows sample number -> PART NO (from UI)
                self.file_select_combo.addItem(name)
                sample_idx = len(self.samples)
                # Build an item widget with checkbox + color square + label so user can select and reorder
                item_w = QWidget()
                item_h = QHBoxLayout()
                item_h.setContentsMargins(2, 2, 2, 2)
                color_lbl = QLabel(f"<span style='color:{color}'>■</span>")
                item_h.addWidget(color_lbl)
                txt_lbl = QLabel(f"{sample_idx}: {part_no_val}")
                txt_lbl.mousePressEvent = lambda _event, n=name: self.set_file_selection_by_name(n)
                txt_lbl.setMaximumWidth(350)
                txt_lbl.setStyleSheet("font-size:11px;")
                item_h.addWidget(txt_lbl)
                # per-item delete button
                del_btn = QPushButton("✖")
                del_btn.setFixedSize(22, 22)
                del_btn.setToolTip("Delete this imported file")
                del_btn.setStyleSheet("""
                    QPushButton {
                        background-color: transparent;
                        color: #9e9e9e;
                        font-family: Arial;
                        font-weight: bold;
                        font-size: 16px;
                        border-radius: 10px;
                        border: none;
                        padding: 0px;
                        margin: 0px;
                    }
                    QPushButton:hover {
                        background-color: #ffebee;
                        color: #f44336;
                    }
                """)
                # connect with closure capturing the widget
                del_btn.clicked.connect(lambda _checked, w=item_w: self.delete_legend_item(w))
                item_h.addWidget(del_btn)
                # item_h.addStretch()
                item_w.setLayout(item_h)
                item_w.setFixedHeight(28)
                self.files_legend_layout.addWidget(item_w)
                self.files_legend_items.append({'widget': item_w, 'color': color_lbl, 'label': txt_lbl, 'delete': del_btn})
                added += 1
                self.file_label.setText(name)
            except Exception:
                continue

        if added > 0:
            self.samples_label.setText(f"Samples: {len(self.samples)}")
            # adjust selector ranges and update plot/avg
            try:
                self.on_file_selection_changed(self.file_select_combo.currentIndex())
                self.update_plot()
                QApplication.processEvents()
                self.update_average()
            except Exception:
                pass
            try:
                pass
            except Exception:
                pass
            QMessageBox.information(self, "Added", f"Added {added} sample(s).")

    def add_files(self):
        """Add multiple CSV files as samples (does not replace current view)."""
        paths, _ = QFileDialog.getOpenFileNames(self, "Add CSV Files", "", "CSV Files (*.csv);;All Files (*)")
        if not paths:
            return

        added = 0
        for p in paths:
            try:
                tlist, trlist, alist, clist = self.read_csv_arrays(p)
                if not tlist:
                    continue

                # Check Mode Restrictions
                mode = self.plot_mode_combo.currentText()
                is_angle = (mode == "Angle vs Torque")
                if is_angle:
                     angle_range = max(alist) - min(alist) if alist else 0
                     if angle_range == 0.0 and all(a == 0.0 for a in alist):
                         continue

                name = os.path.basename(p)
                # PART NO must come from the UI; if empty, keep empty string
                part_no_val = self.part_no_edit.text().strip()
                sample = {
                    'path': p, 
                    'name': name, 
                    'time': tlist, 
                    'torque': trlist, 
                    'angle': alist, 
                    'cycle': clist, 
                    'selected_cycles': sorted(list(set(clist))) if clist else [1],
                    'part_no': part_no_val
                }
                self.samples.append(sample)
                # Add to selector and legend; legend shows sample number -> PART NO
                self.file_select_combo.addItem(name)
                color = self._colors[(len(self.samples)-1) % len(self._colors)]
                sample_idx = len(self.samples)
                item_w = QWidget()
                item_h = QHBoxLayout()
                item_h.setContentsMargins(2, 2, 2, 2)
                color_lbl = QLabel(f"<span style='color:{color}'>■</span>")
                item_h.addWidget(color_lbl)
                txt_lbl = QLabel(f"{sample_idx}: {part_no_val}")
                txt_lbl.mousePressEvent = lambda _event, n=name: self.set_file_selection_by_name(n)
                txt_lbl.setMaximumWidth(350)
                txt_lbl.setStyleSheet("font-size:11px;")
                item_h.addWidget(txt_lbl)
                # per-item delete button
                del_btn = QPushButton("x")
                del_btn.setFixedSize(20, 20)
                del_btn.setToolTip("Delete this imported file")
                del_btn.setStyleSheet("""
                    QPushButton {
                        background-color: transparent;
                        color: #9e9e9e;
                        font-family: Arial;
                        font-weight: bold;
                        font-size: 16px;
                        border-radius: 10px;
                        border: none;
                        padding: 0px;
                        margin: 0px;
                    }
                    QPushButton:hover {
                        background-color: #ffebee;
                        color: #f44336;
                    }
                """)
                del_btn.clicked.connect(lambda _checked, w=item_w: self.delete_legend_item(w))
                item_h.addWidget(del_btn)
                # item_h.addStretch()
                item_w.setLayout(item_h)
                item_w.setFixedHeight(28)
                self.files_legend_layout.addWidget(item_w)
                self.files_legend_items.append({'widget': item_w, 'color': color_lbl, 'label': txt_lbl, 'delete': del_btn})
                added += 1
            except Exception:
                continue

        if added > 0:
            self.samples_label.setText(f"Samples: {len(self.samples)}")
            QMessageBox.information(self, "Added", f"Added {added} sample(s).")
            # adjust selector ranges if 'All' currently selected
            try:
                self.on_file_selection_changed(self.file_select_combo.currentIndex())
            except Exception:
                pass
            pass

    def load_file_from_path_signal(self, path):
        """Slot to handle import from signal."""
        self.load_file_from_path(path)

    def load_file_from_path(self, path):
        """Add a CSV file from path as a sample."""
        if not path or not os.path.exists(path):
            return False

        try:
            try:
                # Attempt read - expected 4 values: time, torque, angle, cycle
                tlist, trlist, alist, clist = self.read_csv_arrays(path)
            except Exception:
                tlist = []
            
            if not tlist:
                return False
            
            # Check Mode Restrictions
            mode = self.plot_mode_combo.currentText()
            is_angle = (mode == "Angle vs Torque")
            
            if is_angle:
                # Check if angle data is present
                angle_range = max(alist) - min(alist) if alist else 0
                if angle_range == 0.0 and all(a == 0.0 for a in alist):
                     QMessageBox.warning(self, "Invalid File", "File does not contain Angle data required for Angle vs Torque mode.\nPlease use 'May chuan' format or switch mode.")
                     return False

            name = os.path.basename(path)
            # PART NO must come from the UI; if empty, keep empty string
            part_no_val = self.part_no_edit.text().strip() if hasattr(self, 'part_no_edit') else ""
            sample = {
                'path': path, 
                'name': name, 
                'time': tlist, 
                'torque': trlist, 
                'angle': alist, 
                'cycle': clist,
                'selected_cycles': sorted(list(set(clist))) if clist else [1],
                'part_no': part_no_val
            }
            self.samples.append(sample)
            # Add to selector and legend
            self.file_select_combo.addItem(name)
            # Keep selection on 'All' (index 0) or update correctly
            self.file_select_combo.setCurrentIndex(0)
            
            color = self._colors[(len(self.samples)-1) % len(self._colors)]
            sample_idx = len(self.samples)
            
            item_w = QWidget()
            item_h = QHBoxLayout()
            item_h.setContentsMargins(2, 2, 2, 2)
            color_lbl = QLabel(f"<span style='color:{color}'>■</span>")
            item_h.addWidget(color_lbl)
            txt_lbl = QLabel(f"{sample_idx}: {part_no_val}")
            txt_lbl.mousePressEvent = lambda _event, n=name: self.set_file_selection_by_name(n)
            txt_lbl.setMaximumWidth(350)
            txt_lbl.setStyleSheet("font-size:11px;")
            item_h.addWidget(txt_lbl)
            
            del_btn = QPushButton("x")
            del_btn.setFixedSize(20, 20)
            del_btn.setToolTip("Delete this imported file")
            del_btn.setStyleSheet("QPushButton { background-color: transparent; color: #9e9e9e; font-weight: bold; border: none; } QPushButton:hover { color: red; }")
            del_btn.clicked.connect(lambda _checked, w=item_w: self.delete_legend_item(w))
            item_h.addWidget(del_btn)
            
            item_w.setLayout(item_h)
            item_w.setFixedHeight(28)
            self.files_legend_layout.addWidget(item_w)
            self.files_legend_items.append({'widget': item_w, 'color': color_lbl, 'label': txt_lbl, 'delete': del_btn})
            
            # Renumber and update cycle text on all labels
            self._rebuild_legend_labels()
            
            self.samples_label.setText(f"Samples: {len(self.samples)}")
            QMessageBox.information(self, "Imported", f"Imported file: {name}")
            
            # Update View
            self.on_file_selection_changed(self.file_select_combo.currentIndex())
            self.update_plot()
            return True
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to load file: {e}")
            return False

    def read_csv_arrays(self, path: str):
        """Read CSV and return (time_list, torque_list, angle_list, cycle_list)."""
        time_list = []
        torque_list = []
        angle_list = []
        cycle_list = []
        with open(path, 'r', encoding='utf-8', errors='ignore') as f:
            lines = [ln.strip() for ln in f if ln.strip()]

        ctr_entries = []
        if any(ln == "END_OF_HEADER" for ln in lines):
            in_data = False
            for line in lines:
                if line == "END_OF_HEADER":
                    in_data = True
                    continue
                if not in_data:
                    continue
                parts = line.split(',')
                # Headers: [Save, State, Cycle, Time, Command, Angle, Torque]
                if len(parts) >= 7:
                    try:
                        # Cycle can be formatted as float like '1.000000', use float() first
                        cycle_val = int(float(parts[2]))
                        time_val = float(parts[3])
                        angle_val = float(parts[5])
                        torque_val = float(parts[6])
                        ctr_entries.append((time_val, torque_val, angle_val, cycle_val))
                    except ValueError:
                        continue
        
        if ctr_entries:
            for t, tr, a, c in ctr_entries:
                time_list.append(t)
                torque_list.append(tr)
                angle_list.append(a)
                cycle_list.append(c)
            return time_list, torque_list, angle_list, cycle_list

        # Try simple two-column (Time, Torque) - no angle, no cycle
        for line in lines:
            parts = line.split(',')
            if len(parts) >= 2:
                try:
                    time_val = float(parts[0])
                    torque_val = float(parts[1])
                    time_list.append(time_val)
                    torque_list.append(torque_val)
                    angle_list.append(0.0) # Default angle
                    cycle_list.append(1)   # Default cycle
                except ValueError:
                    continue
        
        return time_list, torque_list, angle_list, cycle_list

    def on_plot_mode_changed(self, index=None):
        """Handle switch between Time vs Torque and Angle vs Torque."""
        # Update UI Labels
        mode = self.plot_mode_combo.currentText()
        is_angle = (mode == "Angle vs Torque")
        
        # Update Plot Group Box Title
        if getattr(self, 'plot_group', None):
            self.plot_group.setTitle(f"📈 {mode} Plot")
        
        # Update Manual Range Labels/Suffix
        if getattr(self, 'start_label_widget', None):
            self.start_label_widget.setText("Angle Min:" if is_angle else "Start Time (s):")
        if getattr(self, 'end_label_widget', None):
            self.end_label_widget.setText("Angle Max:" if is_angle else "End Time (s):")
            
        if getattr(self, 'start_time_spin', None):
            self.start_time_spin.setSuffix(" deg" if is_angle else " s")
        if getattr(self, 'end_time_spin', None):
            self.end_time_spin.setSuffix(" deg" if is_angle else " s")
        
        # Update Plot Axis Labels and Title immediately
        if getattr(self, 'ax', None):
            if is_angle:
                self.ax.set_xlabel("Angle (deg)", fontsize=10)
                self.ax.set_title("Torque vs Angle", fontsize=12, fontweight='bold')
            else:
                self.ax.set_xlabel("Time (s)", fontsize=10)
                self.ax.set_title("Torque vs Time", fontsize=12, fontweight='bold')
            
            # Redraw canvas to show changes
            try:
                self.canvas.draw_idle()
            except:
                try:
                    self.canvas.draw()
                except:
                    pass
        
        # Trigger update of ranges/plot with data if available
        try:
            self.on_part_name_changed()
        except:
            self.update_plot()



    def on_file_selection_changed(self, index: int):
        """Update STT ranges and cycle list when selected file changes."""
        try:
            sel = self.file_select_combo.currentText()
        except Exception:
            sel = 'All'

        if sel == 'All':
            # choose max length among samples (or single dataset)
            maxlen = 0
            if self.samples:
                maxlen = max((len(s.get('time', [])) for s in self.samples), default=0)
            else:
                maxlen = len(self.time_data)
            if maxlen <= 0:
                maxlen = 1
            # set index-based spinboxes
            self.start_spin.setRange(1, maxlen)
            self.end_spin.setRange(1, maxlen)
            self.start_spin.setValue(1)
            self.end_spin.setValue(maxlen)
            # if range mode is Default and part config exists, sync time spins
            try:
                if getattr(self, 'range_mode_combo', None) and self.range_mode_combo.currentText() == 'Default':
                    # if a part_name is selected, apply its configured times
                    pname = self.part_name_combo.currentText()
                    item_name = self.test_item_combo.currentText()
                    is_angle = (self.plot_mode_combo.currentText() == "Angle vs Torque")
                    ranges = self.test_item_angle_ranges if is_angle else self.test_item_time_ranges
                    pr = ranges.get(pname, {}).get(item_name, {})
                    if pr and self.start_time_spin and self.end_time_spin:
                        self.start_time_spin.setValue(float(pr.get('start', 0.0)))
                        self.end_time_spin.setValue(float(pr.get('end', 0.0)))
            except Exception:
                pass
        else:
            # find selected sample
            s = next((x for x in self.samples if x['name'] == sel), None)
            if s:
                n = len(s.get('time', []))
                if n <= 0:
                    n = 1
                self.start_spin.setRange(1, n)
                self.end_spin.setRange(1, n)
                self.start_spin.setValue(1)
                self.end_spin.setValue(n)
                # if Default mode and a part config exists for selected part name, apply its times
                try:
                    if getattr(self, 'range_mode_combo', None) and self.range_mode_combo.currentText() == 'Default':
                        pname = self.part_name_combo.currentText()
                        item_name = self.test_item_combo.currentText()
                        is_angle = (self.plot_mode_combo.currentText() == "Angle vs Torque")
                        ranges = self.test_item_angle_ranges if is_angle else self.test_item_time_ranges
                        pr = ranges.get(pname, {}).get(item_name, {})
                        if pr and self.start_time_spin and self.end_time_spin:
                            self.start_time_spin.setValue(float(pr.get('start', 0.0)))
                            self.end_time_spin.setValue(float(pr.get('end', 0.0)))
                except Exception:
                    pass
        
        # Refresh cycle checkboxes
        self.refresh_cycle_list()
        
        # update average/plot
        try:
            self.update_average()
        except Exception:
            pass

    def refresh_cycle_list(self):
        """Dynamically create checkboxes for cycles present in selected file(s)."""
        # Block signals to avoid triggering updates while building
        for cb in self.cycle_checkboxes:
            cb.blockSignals(True)
            cb.setParent(None)
            cb.deleteLater()
        self.cycle_checkboxes = []

        sel = self.file_select_combo.currentText()
        all_possible_cycles = set()
        for s in self.samples:
            all_possible_cycles.update(s.get('cycle', []))
        
        if not all_possible_cycles and (self.samples or self.time_data):
            all_possible_cycles = {1}
            
        target_s = None
        if sel != 'All':
            target_s = next((x for x in self.samples if x['name'] == sel), None)
            visible_cycles = sorted(list(set(target_s.get('cycle', [])))) if target_s else sorted(list(all_possible_cycles))
        else:
            visible_cycles = sorted(list(all_possible_cycles))

        from PyQt5.QtWidgets import QCheckBox
        for c in visible_cycles:
            cb = QCheckBox(f"C{c}")
            # Determine initial state
            if target_s:
                is_checked = (c in target_s.get('selected_cycles', []))
            else:
                # In 'All' mode, check it if ANY sample has it selected
                is_checked = any(c in s.get('selected_cycles', []) for s in self.samples)
            
            cb.setChecked(is_checked)
            cb.setProperty('cycle', c)
            cb.setStyleSheet("QCheckBox { font-size: 11px; margin-right: 5px; }")
            cb.stateChanged.connect(self.on_cycle_toggled)
            self.cycle_layout.addWidget(cb)
            self.cycle_checkboxes.append(cb)

    def on_cycle_toggled(self, state):
        """Save selected cycles back to samples and refresh display."""
        try:
            sel = self.file_select_combo.currentText()
            # If nothing imported, skip
            if not self.samples:
                return

            checked_nums = [cb.property('cycle') for cb in self.cycle_checkboxes if cb.isChecked()]
            
            if sel == 'All':
                # Global adjustment: apply filter to all samples
                for s in self.samples:
                    possible = set(s.get('cycle', []))
                    if not possible: possible = {1}
                    s['selected_cycles'] = [c for c in checked_nums if c in possible]
            else:
                # Individual file logic
                s_idx = -1
                for i, x in enumerate(self.samples):
                    if x['name'] == sel:
                        s_idx = i
                        break
                
                if s_idx != -1:
                    s = self.samples[s_idx]
                    if not checked_nums:
                        # User deselected everything for this file
                        from PyQt5.QtWidgets import QMessageBox
                        msg = QMessageBox(self)
                        msg.setWindowTitle("Xác nhận gỡ bỏ tệp tin")
                        msg.setIcon(QMessageBox.Question)
                        msg.setText(f"Hiện tại tất cả các chu kỳ của tệp tin đã bị bỏ chọn.")
                        msg.setInformativeText(f"Bạn có muốn gỡ bỏ hoàn toàn tệp <b>{sel}</b> khỏi ứng dụng không?\n\n"
                                             "• Nhấn 'Xóa tệp' để xác nhận gỡ bỏ.\n"
                                             "• Nhấn 'Hủy bỏ' để giữ lại file (tự động chọn lại 1 chu kỳ).")
                        
                        btn_delete = msg.addButton("Xóa tệp", QMessageBox.AcceptRole)
                        btn_cancel = msg.addButton("Hủy bỏ", QMessageBox.RejectRole)
                        msg.setDefaultButton(btn_cancel)
                        msg.exec_()
                        
                        if msg.clickedButton() == btn_delete:
                            # Delete without sub-confirmation
                            widget = self.files_legend_items[s_idx]['widget']
                            self.delete_legend_item(widget, confirm=False)
                            return
                        else:
                            # Re-check at least one. Try to re-check the sender if it was the one unchecked.
                            sender = self.sender()
                            from PyQt5.QtWidgets import QCheckBox
                            if sender and isinstance(sender, QCheckBox):
                                sender.blockSignals(True)
                                sender.setChecked(True)
                                sender.blockSignals(False)
                            elif self.cycle_checkboxes:
                                self.cycle_checkboxes[0].blockSignals(True)
                                self.cycle_checkboxes[0].setChecked(True)
                                self.cycle_checkboxes[0].blockSignals(False)
                            
                            # Re-calculate checked_nums after restoring one
                            checked_nums = [cb.property('cycle') for cb in self.cycle_checkboxes if cb.isChecked()]

                    s['selected_cycles'] = checked_nums

            self._rebuild_legend_labels()
            self.update_average()
            self.update_plot()
        except Exception:
            pass

    def select_all_cycles(self):
        for cb in self.cycle_checkboxes:
            cb.blockSignals(True)
            cb.setChecked(True)
            cb.blockSignals(False)
        self.on_cycle_toggled(None)

    def deselect_all_cycles(self):
        for cb in self.cycle_checkboxes:
            cb.blockSignals(True)
            cb.setChecked(False)
            cb.blockSignals(False)
        self.on_cycle_toggled(None)

    def open_part_range_setup(self):
        """Open the Part/Test Item Range configuration dialog."""
        parts = [self.part_name_combo.itemText(i) for i in range(self.part_name_combo.count())]
        test_items = [self.test_item_combo.itemText(i) for i in range(self.test_item_combo.count())]
        
        mode = self.plot_mode_combo.currentText()
        is_angle = (mode == "Angle vs Torque")
        
        ranges = self.test_item_angle_ranges if is_angle else self.test_item_time_ranges
        title = "Test Item Angle Range setup" if is_angle else "Test Item Time Range setup"
        labels = ("Start (deg)", "End (deg)") if is_angle else ("Start (s)", "End (s)")
        
        dlg = TestItemRangeDialog(self, parts, test_items, ranges, title=title, labels=labels)
        
        if dlg.exec_() == QDialog.Accepted:
            if is_angle:
                self.test_item_angle_ranges = dlg.ranges or {}
                # Save Angle Ranges
                try:
                    cfg = get_config_file('test_item_angle_ranges.json')
                    with open(cfg, 'w', encoding='utf-8') as f:
                        json.dump(self.test_item_angle_ranges, f, indent=2)
                except: pass
                QMessageBox.information(self, 'Saved', 'Saved Test Item Angle Ranges')
            else:
                self.test_item_time_ranges = dlg.ranges or {}
                # Save Time Ranges
                try:
                    cfg = get_config_file('test_item_time_ranges.json')
                    with open(cfg, 'w', encoding='utf-8') as f:
                        json.dump(self.test_item_time_ranges, f, indent=2)
                except: pass
                QMessageBox.information(self, 'Saved', 'Saved Test Item Time Ranges')
            
            # Refresh UI
            try:
                self.on_part_name_changed()
            except: pass

    def _save_test_item_specs(self):
        """Persist per-part, per-test-item specification settings."""
        cfg = get_config_file('test_item_specs_v2.json')
        with open(cfg, 'w', encoding='utf-8') as f:
            json.dump(self.test_item_specs, f, indent=2)

    def save_current_spec(self):
        """Save the currently visible Spec Min/Max immediately after editing."""
        try:
            if not (getattr(self, 'spec_min_spin', None) and getattr(self, 'spec_max_spin', None)):
                return
            part_name = self.part_name_combo.currentText()
            item_name = self.test_item_combo.currentText()
            if not part_name or not item_name:
                return
            part_specs = self.test_item_specs.setdefault(part_name, {})
            part_specs[item_name] = {
                'min': float(self.spec_min_spin.value()),
                'max': float(self.spec_max_spin.value()),
            }
            self._save_test_item_specs()
        except Exception as exc:
            QMessageBox.warning(self, 'Spec Save Failed', f'Could not save specification:\n{exc}')

    def open_test_item_spec_setup(self):
        """Open the Test Item Specification configuration dialog."""
        test_items = [self.test_item_combo.itemText(i) for i in range(self.test_item_combo.count())]
        parts = [self.part_name_combo.itemText(i) for i in range(self.part_name_combo.count())]
        dlg = TestItemSpecDialog(self, parts, test_items, self.test_item_specs)
        if dlg.exec_() == QDialog.Accepted:
            self.test_item_specs = dlg.specs or {}
            try:
                self._save_test_item_specs()
            except Exception as exc:
                QMessageBox.warning(self, 'Spec Save Failed', f'Could not save specification:\n{exc}')
            self.on_part_name_changed()

    def open_calibration_setup(self):
        """Open the Calibration (Factor K) configuration dialog."""
        # Reload calibration data from file to ensure we have the latest state
        try:
            cfg = get_config_file('calibration.json')
            if cfg.exists():
                with open(cfg, 'r', encoding='utf-8') as f:
                    self.calibration_data = json.load(f)
        except Exception:
            pass # keep existing if fail

        parts = [self.part_name_combo.itemText(i) for i in range(self.part_name_combo.count())]
        dlg = CalibrationDialog(self, parts, self.calibration_data)
        if dlg.exec_() == QDialog.Accepted:
            # reload factor for current part
            self.on_part_name_changed()


    def on_test_item_changed(self, index=None):
        """Load spec min/max when Test Item changes (dependent on Part Name)."""
        try:
            item_name = self.test_item_combo.currentText()
            part_name = self.part_name_combo.currentText()
            
            # Structure: {part: {item: {min, max}}}
            part_specs = self.test_item_specs.get(part_name, {})
            spec = part_specs.get(item_name, {})
            
            if getattr(self, 'spec_min_spin', None) and getattr(self, 'spec_max_spin', None):
                try:
                    self.spec_min_spin.setValue(float(spec.get('min', 0.0)))
                    self.spec_max_spin.setValue(float(spec.get('max', 0.0)))
                except: pass

            # Update Ranges and Refresh Plot if in Default mode
            if getattr(self, 'range_mode_combo', None) and self.range_mode_combo.currentText() == 'Default':
                plot_mode = self.plot_mode_combo.currentText()
                is_angle = (plot_mode == "Angle vs Torque")
                
                ranges = self.test_item_angle_ranges if is_angle else self.test_item_time_ranges
                part_ranges = ranges.get(part_name, {})
                pr = part_ranges.get(item_name, {})
                
                if pr:
                    try:
                        if not is_angle and getattr(self, 'start_time_spin', None) and getattr(self, 'end_time_spin', None):
                            self.start_time_spin.setValue(float(pr.get('start', 0.0)))
                            self.end_time_spin.setValue(float(pr.get('end', 0.0)))
                    except Exception: pass

            # Refresh averages/plot
            try:
                self.update_average()
            except Exception:
                try:
                    self.update_plot()
                except Exception: pass
        except Exception:
            pass

    def on_test_purpose_changed(self, index=None):
        """Show/hide custom text input for Test Purpose and popup if Others."""
        try:
            txt = self.test_purpose_combo.currentText()
            if txt == "Others":
                current_val = self.test_purpose_other_edit.text()
                text, ok = QInputDialog.getText(self, "Test Purpose", "Enter Other Purpose:", text=current_val)
                if ok:
                     self.test_purpose_other_edit.setText(text)
                self.test_purpose_other_edit.show()
            else:
                self.test_purpose_other_edit.hide()
                self.test_purpose_other_edit.clear()
        except Exception:
            pass

    def on_range_mode_changed(self, index=None):
        """Handle range mode switch between Default and Manual.

        - Show manual time spinboxes when Manual is selected.
        - Hide them when Default is selected and apply saved part ranges if present.
        """
        try:
            mode = self.range_mode_combo.currentText() if getattr(self, 'range_mode_combo', None) else 'Default'
        except Exception:
            mode = 'Default'

        # If manual widgets are present, show/hide them
        try:
            if self.start_time_spin and self.end_time_spin:
                show_manual = (mode == 'Manual')
                
                self.start_time_spin.setVisible(show_manual)
                self.end_time_spin.setVisible(show_manual)
                
                if getattr(self, 'start_label_widget', None):
                    self.start_label_widget.setVisible(show_manual)
                if getattr(self, 'end_label_widget', None):
                    self.end_label_widget.setVisible(show_manual)
        except Exception:
            pass

        # If switching to Default, and there is a configured range for current part/item, apply it
        try:
            if mode == 'Default' and getattr(self, 'part_name_combo', None):
                pname = self.part_name_combo.currentText()
                item_name = self.test_item_combo.currentText()
                plot_mode = self.plot_mode_combo.currentText()
                is_angle = (plot_mode == "Angle vs Torque")
                
                ranges = self.test_item_angle_ranges if is_angle else self.test_item_time_ranges
                part_ranges = ranges.get(pname, {})
                pr = part_ranges.get(item_name, {})
                
                if pr:
                    try:
                        # Only populate the time spin controls when in Time mode.
                        if not is_angle and getattr(self, 'start_time_spin', None) and getattr(self, 'end_time_spin', None):
                            self.start_time_spin.setValue(float(pr.get('start', 0.0)))
                            self.end_time_spin.setValue(float(pr.get('end', 0.0)))
                    except Exception:
                        pass

        except Exception:
            pass

        # Refresh averages/plot to reflect new mode
        try:
            self.update_average()
        except Exception:
            try:
                self.update_plot()
            except Exception:
                pass

    def on_part_name_changed(self, index=None):
        """Apply per-part configured range and available specs when part selection changes."""
        try:
            pname = self.part_name_combo.currentText()
            
            # 1. Factor K (Calibration) - Update first so it's ready for plot refreshes
            try:
                # Use stripped part name key
                self.k_factor = float(self.calibration_data.get(pname.strip(), 1.0))
            except:
                self.k_factor = 1.0

            # 2. Specs and Ranges - Trigger unified update in on_test_item_changed
            try:
                self.on_test_item_changed()
            except: pass
        except Exception:
            pass
            
    def get_current_range_values(self, sample):
        """Get torque values for the given sample based on current range settings."""
        import numpy as np
        times = np.array(sample.get('time', []))
        trqs  = np.array(sample.get('torque', []))
        angles = np.array(sample.get('angle', []))
        cycles = np.array(sample.get('cycle', []))
        
        if len(times) == 0 or len(trqs) == 0:
            return []

        # 1. Filter by Cycle Selection
        selected_c = sample.get('selected_cycles', [])
        if len(cycles) > 0 and selected_c:
            c_mask = np.isin(cycles, selected_c)
            times = times[c_mask]
            trqs = trqs[c_mask]
            if len(angles) > 0:
                angles = angles[c_mask]
        
        if len(trqs) == 0:
            return []

        # Determine whether current plot mode is Angle or Time
        try:
            mode_text = (self.plot_mode_combo.currentText() or '').strip().lower() if getattr(self, 'plot_mode_combo', None) else ''
            is_angle_mode = mode_text.startswith('angle')
        except Exception:
            is_angle_mode = False

        # Determine range selection mode (Default/Manual)
        range_mode = 'Default'
        try:
            if getattr(self, 'range_mode_combo', None):
                range_mode = self.range_mode_combo.currentText()
        except: pass

        # Determine start/end values depending on mode (treat values as Angle when in Angle mode)
        st = None
        en = None

        try:
            pname = self.part_name_combo.currentText() if getattr(self, 'part_name_combo', None) else None
        except Exception:
            pname = None

        if range_mode == 'Manual':
            # Manual: use the start/end spin controls as numeric bounds.
            # If in Angle mode, interpret these as angle bounds; otherwise as time bounds.
            if getattr(self, 'start_time_spin', None) and getattr(self, 'end_time_spin', None):
                try:
                    st = float(self.start_time_spin.value())
                    en = float(self.end_time_spin.value())
                except Exception:
                    st = en = None
        else:
            # Default mode must NOT silently crop imported cycle data.
            # The Test Item default ranges are report/config references; using them here
            # made Data Info show only a small slice (e.g. C3 = 33 pts) even though
            # the CSV contains the full cycle (e.g. C3 = 869 pts). Keep the full
            # selected cycle(s) unless the user explicitly switches to Manual range.
            st = en = None

        # If we have valid numeric bounds, select by the appropriate X-axis (angle or time)
        if st is not None and en is not None:
            try:
                # Choose x-axis values: angle array when in Angle mode, otherwise times
                if is_angle_mode:
                    x_vals = angles if len(angles) > 0 else times
                else:
                    x_vals = times

                # If no x_vals, fall back to index-based selection later
                if len(x_vals) > 0:
                    # Build list of indices where x is within [st,en]
                    idxs = []
                    for i, xv in enumerate(x_vals):
                        try:
                            if xv is None:
                                continue
                            xv_f = float(xv)
                            if xv_f >= st and xv_f <= en:
                                idxs.append(i)
                        except Exception:
                            continue

                    if idxs:
                        si = idxs[0]
                        ei = idxs[-1] + 1
                        # clamp
                        si = max(0, min(si, len(trqs)-1))
                        ei = max(si+1, min(ei, len(trqs)))
                        sub_trqs = trqs[si:ei]
                        # Apply K factor
                        try:
                            k = self.k_factor
                            if k != 1.0:
                                sub_trqs = [t * k for t in sub_trqs]
                        except: pass
                        return sub_trqs
            except Exception:
                pass
        
        # Fallback to current spin indices if no time range 
        # (or for Default mode when no Part Range configured)
        try:
            start_idx = max(0, min(self.start_spin.value() - 1, len(trqs)-1))
            end_idx = max(start_idx + 1, min(self.end_spin.value(), len(trqs)))
            sub_trqs = trqs[start_idx:end_idx]
            
            # Apply K factor
            try:
                k = self.k_factor
                if k != 1.0:
                    sub_trqs = [t * k for t in sub_trqs]
            except: pass
            
            return sub_trqs
        except:
            return trqs

    def remove_selected(self):
        """Remove the currently selected imported sample (not 'All')."""
        try:
            sel = self.file_select_combo.currentText()
        except Exception:
            sel = 'All'
        if sel == 'All':
            QMessageBox.information(self, "Remove", "Please select a specific file to remove (not 'All').")
            return
        idx = self.file_select_combo.currentIndex()
        if idx <= 0:
            QMessageBox.information(self, "Remove", "No file selected.")
            return
        # Confirm
        ok = QMessageBox.question(self, "Confirm Remove", f"Remove '{sel}' from imported samples?", QMessageBox.Yes | QMessageBox.No)
        if ok != QMessageBox.Yes:
            return
        sample_idx = idx - 1
        try:
            # remove sample data
            del self.samples[sample_idx]
        except Exception:
            pass
        # remove combo item
        try:
            self.file_select_combo.removeItem(idx)
        except Exception:
            pass
        # remove legend item widget
        try:
            item = self.files_legend_items.pop(sample_idx)
            w = item.get('widget')
            self.files_legend_layout.removeWidget(w)
            w.deleteLater()
        except Exception:
            pass
        # renumber remaining legend labels
        try:
            for i, item in enumerate(self.files_legend_items, start=1):
                try:
                    lbl = item.get('label')
                    lbl.setText(f"{i}: " + lbl.text().split(':', 1)[1].strip() if ':' in lbl.text() else f"{i}: {lbl.text()}")
                except Exception:
                    pass
        except Exception:
            pass

        # update UI counts and plot
        try:
            self.samples_label.setText(f"Samples: {len(self.samples)}")
            # adjust selector ranges and update plot/avg
            self.on_file_selection_changed(self.file_select_combo.currentIndex())
            self.update_plot()
            QApplication.processEvents()
            self.update_average()
            pass
        except Exception:
            pass

    def select_all(self):
        # removed: global select not applicable after checkbox removal
        return

    def select_none(self):
        # removed: global deselect not applicable after checkbox removal
        return
    # select/toggle UI removed: per-item delete buttons used and checkboxes control inclusion

    def move_up(self):
        """Move currently selected combo file up by one position."""
        idx = self.file_select_combo.currentIndex()
        if idx <= 1:
            return
        i = idx - 1
        # swap samples
        try:
            self.samples[i-1], self.samples[i] = self.samples[i], self.samples[i-1]
        except Exception:
            pass
        # swap combo items
        try:
            text_i = self.file_select_combo.itemText(idx)
            text_j = self.file_select_combo.itemText(idx-1)
            self.file_select_combo.setItemText(idx, text_j)
            self.file_select_combo.setItemText(idx-1, text_i)
            self.file_select_combo.setCurrentIndex(idx-1)
        except Exception:
            pass
        # swap legend widgets and update numbering
        try:
            self.files_legend_items[i-1], self.files_legend_items[i] = self.files_legend_items[i], self.files_legend_items[i-1]
            self._rebuild_legend_labels()
        except Exception:
            pass

    def move_down(self):
        """Move currently selected combo file down by one position."""
        idx = self.file_select_combo.currentIndex()
        if idx <= 0 or idx >= self.file_select_combo.count()-1:
            return
        i = idx - 1
        try:
            self.samples[i], self.samples[i+1] = self.samples[i+1], self.samples[i]
        except Exception:
            pass
        try:
            text_i = self.file_select_combo.itemText(idx)
            text_j = self.file_select_combo.itemText(idx+1)
            self.file_select_combo.setItemText(idx, text_j)
            self.file_select_combo.setItemText(idx+1, text_i)
            self.file_select_combo.setCurrentIndex(idx+1)
        except Exception:
            pass
        try:
            self.files_legend_items[i], self.files_legend_items[i+1] = self.files_legend_items[i+1], self.files_legend_items[i]
            self._rebuild_legend_labels()
        except Exception:
            pass

    def _rebuild_legend_labels(self):
        """Update the numeric prefixes and cycle summary on legend item labels."""
        try:
            for i, item in enumerate(self.files_legend_items, start=1):
                try:
                    lbl = item.get('label')
                    sample = self.samples[i-1]
                    part_no = sample.get('part_no', '')
                    sel_c = sample.get('selected_cycles', [])
                    c_str = ", ".join([f"C{c}" for c in sorted(sel_c)])
                    if c_str:
                        text = f"{i}: {part_no} ({c_str})"
                    else:
                        text = f"{i}: {part_no} (None)"
                    lbl.setText(text)
                    lbl.setToolTip(sample.get('name', ''))
                    # Make it look clickable
                    lbl.setCursor(Qt.PointingHandCursor)
                except Exception:
                    pass
        except Exception:
            pass

    def set_file_selection_by_name(self, name):
        """Set the cycle selector combo to the specified file name."""
        try:
            index = self.file_select_combo.findText(name)
            if index >= 0:
                self.file_select_combo.setCurrentIndex(index)
        except Exception:
            pass

    def delete_legend_item(self, widget, confirm=True):
        """Delete a sample corresponding to the legend widget passed in.
        This will prompt for confirmation (if confirm=True) and then remove the sample, its
        combo entry, and the legend widget. It keeps samples and legend lists
        in sync and updates the plot and stats.
        """
        if widget is None:
            return
        # find index
        idx = None
        try:
            for i, item in enumerate(self.files_legend_items):
                if item.get('widget') is widget:
                    idx = i
                    break
        except Exception:
            idx = None
        if idx is None:
            return
        name = self.samples[idx].get('name') if idx < len(self.samples) else 'sample'
        
        if confirm:
            ok = QMessageBox.question(self, "Xác nhận xóa", f"Bạn có chắc chắn muốn gỡ bỏ tệp '{name}'?", QMessageBox.Yes | QMessageBox.No)
            if ok != QMessageBox.Yes:
                return
        try:
            # remove sample data
            del self.samples[idx]
        except Exception:
            pass
        try:
            # remove combo entry (offset by 1 because 'All' at index 0)
            if self.file_select_combo.count() > idx + 1:
                self.file_select_combo.removeItem(idx + 1)
        except Exception:
            pass
        try:
            item = self.files_legend_items.pop(idx)
            w = item.get('widget')
            # remove widget from layout and delete
            try:
                self.files_legend_layout.removeWidget(w)
                w.deleteLater()
            except Exception:
                pass
        except Exception:
            pass
        # renumber and refresh
        try:
            self._rebuild_legend_labels()
            self.samples_label.setText(f"Samples: {len(self.samples)}")
            self.on_file_selection_changed(self.file_select_combo.currentIndex())
            self.update_plot()
            QApplication.processEvents()
            self.update_average()
        except Exception:
            pass

    def clear_all_samples(self):
        """Remove all imported samples after confirmation."""
        if not self.samples:
            QMessageBox.information(self, "Clear All", "No imported samples to clear.")
            return
        ok = QMessageBox.question(self, "Confirm Clear All", "Remove ALL imported samples?", QMessageBox.Yes | QMessageBox.No)
        if ok != QMessageBox.Yes:
            return
        # clear data
        try:
            self.samples.clear()
        except Exception:
            self.samples = []
        # clear combo box items except 'All'
        try:
            while self.file_select_combo.count() > 1:
                self.file_select_combo.removeItem(1)
        except Exception:
            pass
        # remove legend widgets
        try:
            for item in self.files_legend_items:
                try:
                    w = item.get('widget')
                    self.files_legend_layout.removeWidget(w)
                    w.deleteLater()
                except Exception:
                    pass
            self.files_legend_items = []
        except Exception:
            pass
        # update UI
        try:
            self.samples_label.setText("Samples: 0")
            self.file_label.setText("No file loaded")
            self.on_file_selection_changed(self.file_select_combo.currentIndex())
            self.update_plot()
            QApplication.processEvents()
            self.update_average()
            pass
        except Exception:
            pass
    
    def load_csv(self, path: str):
        """Load CSV data (CTR format with header)"""
        self.time_data = []
        self.torque_data = []
        # Read all lines first so we can detect format
        with open(path, 'r', encoding='utf-8', errors='ignore') as f:
            lines = [ln.strip() for ln in f if ln.strip()]

        # Try CTR format (lines after END_OF_HEADER with >=7 columns)
        ctr_entries = []
        if any(ln == "END_OF_HEADER" for ln in lines):
            in_data = False
            for line in lines:
                if line == "END_OF_HEADER":
                    in_data = True
                    continue
                if not in_data:
                    continue
                parts = line.split(',')
                if len(parts) >= 7:
                    try:
                        time_val = float(parts[3])
                        torque_val = float(parts[6])
                        ctr_entries.append((time_val, torque_val))
                    except ValueError:
                        continue

        # If CTR parsing produced entries, use them; otherwise try simple two-column format
        if ctr_entries:
            for t, tr in ctr_entries:
                self.time_data.append(t)
                self.torque_data.append(tr)
        else:
            # Attempt to parse each line as `time,torque` where both are floats
            simple_count = 0
            for line in lines:
                parts = line.split(',')
                if len(parts) >= 2:
                    try:
                        time_val = float(parts[0])
                        torque_val = float(parts[1])
                        self.time_data.append(time_val)
                        self.torque_data.append(torque_val)
                        simple_count += 1
                    except ValueError:
                        # skip lines that don't parse as two floats
                        continue

        # If still empty, fallback: try scanning any lines for numeric pairs anywhere
        if not self.time_data:
            for line in lines:
                parts = [p.strip() for p in line.split(',') if p.strip()]
                for i in range(len(parts)-1):
                    try:
                        t = float(parts[i])
                        tr = float(parts[i+1])
                        self.time_data.append(t)
                        self.torque_data.append(tr)
                        break
                    except ValueError:
                        continue
        
        # Update spin box ranges
        n = len(self.time_data)
        if n > 0:
            self.start_spin.setRange(1, n)
            self.end_spin.setRange(1, n)
            self.start_spin.setValue(1)
            self.end_spin.setValue(n)
            self.total_label.setText(f"Pts: {n}")
            self.total_label.setStyleSheet("font-weight: bold; color: #333333; font-size: 11px;")
            # Show higher precision for torque values
            self.min_label.setText(f"{min(self.torque_data):.6f}")
            self.max_label.setText(f"{max(self.torque_data):.6f}")
    
    def update_plot(self):
        """Update plot with current data"""
        self.ax.clear()
        # If samples present, plot each CHECKED sample with its own color
        if self.samples:
            sel = self.file_select_combo.currentText() if hasattr(self, 'file_select_combo') else 'All'
            # Determine current mode and whether selection is time-based or index-based for plotting
            mode = self.plot_mode_combo.currentText() if getattr(self, 'plot_mode_combo', None) else ''
            is_angle = (mode == "Angle vs Torque")
            use_range = False
            try:
                if getattr(self, 'range_mode_combo', None) and self.range_mode_combo.currentText() == 'Manual':
                    # Manual range is explicit user filtering.
                    if not is_angle and getattr(self, 'start_time_spin', None) is not None:
                        use_range = True
                    elif is_angle and getattr(self, 'start_angle_spin', None) is not None:
                        use_range = True
                elif getattr(self, 'range_mode_combo', None) and self.range_mode_combo.currentText() == 'Default':
                    # Do not auto-crop by Test Item default range while reviewing cycles.
                    # This keeps plotted points consistent with the full selected cycle(s).
                    use_range = False
            except Exception:
                use_range = False

            import numpy as np

            for orig_i, s in enumerate(self.samples):
                times = np.array(s.get('time', []))
                trqs = np.array(s.get('torque', []))
                angles = np.array(s.get('angle', []))
                cycles = np.array(s.get('cycle', []))
                selected_c = s.get('selected_cycles', [])
                
                if len(times) == 0 or len(trqs) == 0:
                    continue
                
                # Filter by Cycle Selection
                if len(cycles) > 0 and selected_c:
                    c_mask = np.isin(cycles, selected_c)
                    times = times[c_mask]
                    trqs = trqs[c_mask]
                    if len(angles) > 0:
                        angles = angles[c_mask]

                if len(times) == 0:
                    continue

                # Determine X-Axis data
                x_data = angles if is_angle else times
                if is_angle and (len(angles) == 0 or len(angles) != len(trqs)):
                    # Strictly no fallback to keep plot consistent with mode
                    x_data = np.array([]) 
                
                color = self._colors[orig_i % len(self._colors)]
                # Determine indices based on selection
                # Only use range filtering if we have start/end spin logic active
                if use_range:
                    try:
                        # Determine st/en based on range mode
                        range_mode = self.range_mode_combo.currentText() if getattr(self, 'range_mode_combo', None) else 'Default'
                        st = None
                        en = None
                        
                        if range_mode == 'Manual' and self.start_time_spin and self.end_time_spin:
                            # Manual mode: use spinboxes
                            st = float(self.start_time_spin.value())
                            en = float(self.end_time_spin.value())
                        elif range_mode == 'Default':
                            # Default mode: do not crop data (matches get_current_range_values)
                            st = None
                            en = None
                        
                        # Only apply mask if we have valid range
                        if st is None or en is None:
                            raise ValueError("No valid range")
                        
                        # Apply numpy mask for filtering
                        import numpy as np
                        arr_x = np.array(x_data)
                        arr_y = np.array(trqs)
                        
                        mask = (arr_x >= st) & (arr_x <= en)
                        
                        plot_x = arr_x[mask]
                        plot_y = arr_y[mask]
                        
                    except Exception:
                         plot_x = x_data
                         plot_y = trqs

                else:
                    if sel == 'All' or sel == s['name']:
                         # No specific range, full data (or index based if that was filtered, but we moved to Filter Logic)
                         # Here we just use full data for simplicity if not in range mode?
                         # The original code had index start/end for 'Default' without parts.
                         start_idx = max(0, min(self.start_spin.value() - 1, len(trqs)-1))
                         end_idx = max(start_idx + 1, min(self.end_spin.value(), len(trqs)))
                         plot_x = x_data[start_idx:end_idx]
                         plot_y = trqs[start_idx:end_idx]
                    else:
                        # plot full series but faded
                        plot_x = x_data
                        plot_y = trqs
                
                # Apply K factor
                try:
                    k = self.k_factor
                    if k != 1.0 and len(plot_y) > 0:
                         plot_y = [t * k for t in plot_y]
                except: pass

                if len(plot_x) > 0 and len(plot_y) > 0:
                    lw = 1.6 if sel == s['name'] else 0.9
                    alpha = 1.0 if (sel == 'All' or sel == s['name']) else 0.5
                    lbl = f"{orig_i+1}"
                    # If Filtering by Angle resulted in gaps, plotting with '-' will define lines across gaps.
                    # It's better to use '.' or accept the lines.
                    self.ax.plot(plot_x, plot_y, marker='.', linestyle='-' , color=color, linewidth=lw, markersize=3, alpha=alpha, label=lbl)

                    # annotate sample number near the last point in the plotted range
                    try:
                        x_annot = plot_x[-1]
                        y_annot = plot_y[-1]
                        self.ax.annotate(f"S{orig_i+1}", xy=(x_annot, y_annot), xytext=(6, 3), textcoords='offset points', color=color, fontsize=9, weight='bold', bbox=dict(facecolor='white', alpha=0.7, edgecolor=color, linewidth=0.5))
                    except Exception:
                        pass
            self.ax.relim()
            self.ax.autoscale_view()
            # Add legend on plot indicating Sample number -> file name
            try:
                leg = self.ax.legend(loc='upper right', fontsize=8, framealpha=0.8)
                leg.get_frame().set_linewidth(0.4)
            except Exception:
                pass
        else:
            # fallback to single loaded dataset
            if len(self.time_data) > 0:
                try:
                    start_idx = self.start_spin.value() - 1
                    end_idx = self.end_spin.value()
                except Exception:
                    start_idx = 0
                    end_idx = len(self.time_data)

                start_idx = max(0, min(start_idx, len(self.time_data) - 1))
                end_idx = max(start_idx + 1, min(end_idx, len(self.time_data)))

                plot_time = self.time_data[start_idx:end_idx]
                plot_torque = self.torque_data[start_idx:end_idx]
                # Apply K factor
                try:
                    k = self.k_factor
                    if k != 1.0 and plot_torque:
                         plot_torque = [t * k for t in plot_torque]
                except: pass
                if plot_time and plot_torque:
                    self.ax.plot(plot_time, plot_torque, 'b.-', linewidth=0.9, markersize=3)
                    self.ax.relim()
                    self.ax.autoscale_view()
        
        # Force Y-axis rules: Ymin=0 if data>=0, else min(data); Ymax=max(data)+margin
        try:
             all_y = []
             for line in self.ax.get_lines():
                 y_data = line.get_ydata()
                 if y_data is not None and len(y_data) > 0:
                     all_y.extend(y_data)
             
             if all_y:
                 d_min = min(all_y)
                 d_max = max(all_y)
                 margin = (d_max - d_min) * 0.1
                 if margin == 0: margin = (d_max * 0.1) if d_max != 0 else 1.0
                 
                 t_max = d_max + margin
                 t_min = 0.0 if d_min >= 0 else d_min
                 self.ax.set_ylim(t_min, t_max)
        except Exception:
             pass

        mode = self.plot_mode_combo.currentText()
        is_angle = (mode == "Angle vs Torque")

        if is_angle:
             self.ax.set_xlabel("Angle (deg)", fontsize=10)
             self.ax.set_title("Torque vs Angle", fontsize=12, fontweight='bold')
        else:
             self.ax.set_xlabel("Time (s)", fontsize=10)
             self.ax.set_title("Torque vs Time", fontsize=12, fontweight='bold')
             
        self.ax.set_ylabel("Torque (N·m)", fontsize=12)
        self.ax.grid(True, alpha=0.3)
        try:
            # Y-axis ticks to match X-axis count for cleaner look
            x_ticks = self.ax.get_xticks()
            n = len(x_ticks)
            self.ax.yaxis.set_major_locator(MaxNLocator(nbins=n))
        except: pass
        try:
            self.fig.tight_layout()
        except Exception:
            pass
        # Use draw_idle to be friendlier with the Qt event loop
        try:
            self.canvas.draw_idle()
        except Exception:
            try:
                self.canvas.draw()
            except Exception:
                pass
    
    def update_average(self):
        """Calculate and display average torque in range"""
        # Track the calculated average for validity judgment
        calculated_average = None

        # If multiple samples loaded, compute average across selected dataset(s)
        if self.samples:
            sel = self.file_select_combo.currentText() if hasattr(self, 'file_select_combo') else 'All'
            total_vals = []
            # Helper to compute subset indices or time-sliced values
            def _slice_vals(vals, times=None):
                if not vals:
                    return []
                # Determine whether to slice by time (or by angle). Manual should
                # apply to time only when in Time mode (or if explicit time controls exist).
                use_time = False
                try:
                    mode_text = (self.plot_mode_combo.currentText() or '').strip().lower() if getattr(self, 'plot_mode_combo', None) else ''
                    is_angle_mode = mode_text.startswith('angle')
                    rm = self.range_mode_combo.currentText() if getattr(self, 'range_mode_combo', None) else 'Default'
                    if rm == 'Manual':
                        # Manual applies to time only when not in Angle mode (unless angle manual controls exist)
                        if not is_angle_mode and getattr(self, 'start_time_spin', None) is not None:
                            use_time = True
                        elif is_angle_mode and getattr(self, 'start_angle_spin', None) is not None:
                            use_time = True
                    elif rm == 'Default':
                        pname = self.part_name_combo.currentText() if hasattr(self, 'part_name_combo') else None
                        item_name = self.test_item_combo.currentText()
                        if pname:
                            ranges = self.test_item_time_ranges if not is_angle_mode else self.test_item_angle_ranges
                            part_ranges = ranges.get(pname, {})
                            if item_name in part_ranges:
                                use_time = True
                except Exception:
                    use_time = False

                if use_time and times is not None and getattr(self, 'start_time_spin', None) and getattr(self, 'end_time_spin', None):
                    try:
                        st = float(self.start_time_spin.value())
                        en = float(self.end_time_spin.value())
                        si = bisect.bisect_left(times, st)
                        ei = bisect.bisect_right(times, en)
                        si = max(0, min(si, len(times)-1))
                        ei = max(si+1, min(ei, len(times)))
                        return vals[si:ei]
                    except Exception:
                        pass
                start_idx = max(0, min(self.start_spin.value() - 1, len(vals)-1))
                end_idx = max(start_idx + 1, min(self.end_spin.value(), len(vals)))
                return vals[start_idx:end_idx]

            if sel == 'All':
                if hasattr(self, 'source_label'):
                    self.source_label.setText("Target: All Samples")
                # apply start/end per file and concatenate across all samples
                import numpy as np
                all_subsets = []
                for i, s in enumerate(self.samples):
                    subset = self.get_current_range_values(s)
                    if len(subset) > 0:
                        all_subsets.append(subset)
                
                if all_subsets:
                    total_vals_arr = np.concatenate(all_subsets)
                    # Convert to absolute values for calculation
                    total_vals_abs = np.abs(total_vals_arr)
                    avg = np.mean(total_vals_abs)
                    calculated_average = avg
                    self.avg_label.setText(f"{avg:.6f}")
                    # Update Data Info labels
                    self.total_label.setText(f"Pts: {len(total_vals_arr)}")
                    min_val = np.min(total_vals_arr)
                    max_val = np.max(total_vals_arr)
                    self.min_label.setText(f"{min_val:.6f}")
                    self.max_label.setText(f"{max_val:.6f}")
                    
                    # Convert to Kgf.cm (1 Nm = 10.1972 Kgf.cm)
                    self.avg_label_kgf.setText(f"{avg * 10.1972:.6f}")
                    self.min_label_kgf.setText(f"{min_val * 10.1972:.6f}")
                    self.max_label_kgf.setText(f"{max_val * 10.1972:.6f}")
                else:
                    self.avg_label.setText("-")
                    self.total_label.setText("Pts: 0")
                    self.min_label.setText("-")
                    self.max_label.setText("-")
                    self.avg_label_kgf.setText("-")
                    self.min_label_kgf.setText("-")
                    self.max_label_kgf.setText("-")
            else:
                # find selected sample
                sel_name = sel
                if hasattr(self, 'source_label'):
                    display_name = sel_name if len(sel_name) < 30 else sel_name[:27] + "..."
                    self.source_label.setText(f"Target: {display_name}")
                s = next((x for x in self.samples if x['name'] == sel_name), None)
                if not s or not s.get('torque'):
                    self.avg_label.setText("-")
                    self.total_label.setText("Pts: 0")
                    self.min_label.setText("-")
                    self.max_label.setText("-")
                else:
                    subset = self.get_current_range_values(s)
                    if len(subset) > 0:
                        import numpy as np
                        # Convert to absolute values for calculation
                        subset_abs = np.abs(subset)
                        avg = np.mean(subset_abs)
                        calculated_average = avg
                        self.avg_label.setText(f"{avg:.6f}")
                        self.total_label.setText(f"Pts: {len(subset)}")
                        min_val = np.min(subset)
                        max_val = np.max(subset)
                        self.min_label.setText(f"{min_val:.6f}")
                        self.max_label.setText(f"{max_val:.6f}")
                        
                        # Convert to Kgf.cm
                        self.avg_label_kgf.setText(f"{avg * 10.1972:.6f}")
                        self.min_label_kgf.setText(f"{min_val * 10.1972:.6f}")
                        self.max_label_kgf.setText(f"{max_val * 10.1972:.6f}")
                    else:
                        self.avg_label.setText("-")
                        self.total_label.setText("Pts: 0")
                        self.min_label.setText("-")
                        self.max_label.setText("-")
                        self.avg_label_kgf.setText("-")
                        self.min_label_kgf.setText("-")
                        self.max_label_kgf.setText("-")
        else:
            # fallback single dataset
            if len(self.torque_data) == 0:
                self.avg_label.setText("-")
                self.judgment_label.setText("")
                return
            start_idx = self.start_spin.value() - 1  # Convert to 0-based
            end_idx = self.end_spin.value()  # end is inclusive
            start_idx = max(0, min(start_idx, len(self.torque_data) - 1))
            end_idx = max(start_idx + 1, min(end_idx, len(self.torque_data)))
            subset = self.torque_data[start_idx:end_idx]
            # Apply K factor
            try:
                k = self.k_factor
                if k != 1.0 and subset:
                    subset = [t * k for t in subset]
            except: pass
            
            if subset:
                # Convert to absolute values for calculation
                subset_abs = [abs(v) for v in subset]
                avg = sum(subset_abs) / len(subset_abs)
                calculated_average = avg
                self.avg_label.setText(f"{avg:.6f}")
                # Update Data Info labels for single dataset
                self.total_label.setText(f"Pts: {len(subset)}")
                min_val = min(subset)
                max_val = max(subset)
                self.min_label.setText(f"{min_val:.6f}")
                self.max_label.setText(f"{max_val:.6f}")
                
                # Convert to Kgf.cm
                self.avg_label_kgf.setText(f"{avg * 10.1972:.6f}")
                self.min_label_kgf.setText(f"{min_val * 10.1972:.6f}")
                self.max_label_kgf.setText(f"{max_val * 10.1972:.6f}")
            else:
                self.avg_label.setText("-")
                self.total_label.setText("Pts: 0")
                self.min_label.setText("-")
                self.max_label.setText("-")
                self.avg_label_kgf.setText("-")
                self.min_label_kgf.setText("-")
                self.max_label_kgf.setText("-")

        # redraw plot to reflect new range selection
        try:
            self.update_plot()
        except Exception:
            pass

        # Judgment logic based directly on calculated average
        try:
            if calculated_average is not None and getattr(self, 'spec_min_spin', None) and getattr(self, 'spec_max_spin', None):
                smin = float(self.spec_min_spin.value())
                smax = float(self.spec_max_spin.value())
                if smin <= calculated_average <= smax:
                    self.judgment_label.setText('OK')
                    self.judgment_label.setStyleSheet('color: green;')
                else:
                    self.judgment_label.setText('NG')
                    self.judgment_label.setStyleSheet('color: red;')
            else:
                self.judgment_label.setText("")
        except Exception:
            pass


    def export_xlsx(self):
        """Export Excel TEST REPORT khổ A4.
        Layout theo sample_report.pdf:
        - Title merge A1:F2
        - Metadata A4:B15
        - TORQUE TYPE & PART TYPE check option ở C-D
        - GRAPH merge D3:I17 có viền đen và chèn ảnh matplotlib
        - Sample table tại row 18
        - GRAPH condition (STT Start/End) sau bảng sample
        - Sheet Data chứa Index-Time-Torque
        """
        # Require either a single-loaded dataset (`time_data`) or imported `samples`
        if not self.time_data and not self.samples:
            QMessageBox.warning(self, "No data", "No data to export. Import or add at least one CSV file.")
            return

        try:
            from openpyxl import Workbook
            from openpyxl.drawing.image import Image as XLImage
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from PIL import Image as PILImage
        except Exception:
            QMessageBox.critical(
                self, "Missing dependency",
                "Please install required packages:\npython -m pip install openpyxl pillow"
            )
            return

        # Default suggested filename: use Report File path directory, not the imported CSV directory.
        timestr = datetime.now().strftime('%Y%m%d_%H%M%S')
        if self.file_path:
            suggested_name = os.path.splitext(os.path.basename(self.file_path))[0] + '.xlsx'
        else:
            suggested_name = f'report_{timestr}.xlsx'
        report_dir = ''
        try:
            report_dir = self.report_path_edit.text().strip()
        except Exception:
            report_dir = getattr(self, 'report_dir', '') or ''
        if not report_dir:
            report_dir = os.path.dirname(self.file_path) if self.file_path else os.getcwd()
        suggested = os.path.join(report_dir, suggested_name)
        path, _ = QFileDialog.getSaveFileName(self, "Save Excel Report", suggested, "Excel Files (*.xlsx)")
        if not path:
            return

        # Render a fixed-size figure to a temporary PNG so export is independent of UI/window size.
        tmp_png = None
        # Target image size in pixels (consistent for all exports)
        # Dynamic ratio based on user input
        ratio = 0.5
        try:
            ratio = self.aspect_ratio_spin.value()
        except: pass
        
        TARGET_W_PX = 1200
        TARGET_H_PX = int(TARGET_W_PX * ratio)
        
        # Recalculate layout rows based on target height
        # Assuming row height ~ 18 points = 24 pixels
        approx_row_h_px = 24
        graph_rows_needed = int(TARGET_H_PX / approx_row_h_px) + 2 # +2 for padding
        
        # Update export constants for this run
        current_graph_bottom = GRAPH_TOP + graph_rows_needed
        current_image_bottom = current_graph_bottom - 2
        # Ensure minimum size
        if current_graph_bottom < GRAPH_TOP + 5:
            current_graph_bottom = GRAPH_TOP + 5
            
        # Choose a DPI and figure size in inches
        # Choose a DPI and figure size in inches
        DPI = 150
        fig_w_in = TARGET_W_PX / DPI
        fig_h_in = TARGET_H_PX / DPI
        fig2 = Figure(figsize=(fig_w_in, fig_h_in), dpi=DPI)
        ax2 = fig2.add_subplot(111)
        
        mode = self.plot_mode_combo.currentText()
        is_angle = (mode == "Angle vs Torque")

        if is_angle:
             ax2.set_xlabel('Angle (deg)', fontsize=10)
             ax2.set_title('Angle vs Torque', fontsize=12, fontweight='bold')
        else:
             ax2.set_xlabel('Time (s)', fontsize=10)
             ax2.set_title('Time vs Torque', fontsize=12, fontweight='bold')

        ax2.set_ylabel('Torque (N·m)', fontsize=10)
        ax2.grid(True, alpha=0.3)
        # Apply Y-axis tick improvement (1/2 of X-ticks)
        try:
             # Force a draw so ticks are calculated
             fig2.canvas.draw()
             n_xticks = len(ax2.get_xticks())
             # Increase Y-axis resolution for Excel export
             ax2.yaxis.set_major_locator(MaxNLocator(nbins=20))
        except Exception:
             pass

        # Prepare Spec Lines Data if available
        spec_min = None
        spec_max = None
        try:
             pname = self.part_name_combo.currentText() if hasattr(self, 'part_name_combo') else None
             item_name = self.test_item_combo.currentText() if hasattr(self, 'test_item_combo') else None
             if pname and item_name:
                 part_specs = self.test_item_specs.get(pname, {})
                 spec = part_specs.get(item_name, {})
                 if spec:
                     spec_min = float(spec.get('min', 0.0))
                     spec_max = float(spec.get('max', 0.0))
        except: pass
        
        # Plot Spec Lines - REMOVED as requested
        # if spec_min is not None:
        #      ax2.axhline(y=spec_min, color='red', linestyle='--', linewidth=1, alpha=0.7, label='Min Spec')
        # if spec_max is not None:
        #      ax2.axhline(y=spec_max, color='red', linestyle='--', linewidth=1, alpha=0.7, label='Max Spec')

        if self.samples:
            sel = self.file_select_combo.currentText() if hasattr(self, 'file_select_combo') else 'All'
            import numpy as np

            for orig_i, s in enumerate(self.samples):
                selected_c = s.get('selected_cycles', [])
                times = np.array(s.get('time', []))
                trqs = np.array(s.get('torque', []))
                angles = np.array(s.get('angle', []))
                cycles = np.array(s.get('cycle', []))

                if len(times) == 0 or len(trqs) == 0:
                    continue
                
                # Filter by Cycle Selection
                if len(cycles) > 0 and selected_c:
                    c_mask = np.isin(cycles, selected_c)
                    times = times[c_mask]
                    trqs = trqs[c_mask]
                    if len(angles) > 0:
                        angles = angles[c_mask]
                
                if len(times) == 0:
                    continue

                # Determine X-Axis data
                x_data = angles if is_angle else times
                if is_angle and (len(angles) == 0 or len(angles) != len(trqs)):
                    # Strictly no fallback to keep plot consistent with mode
                    x_data = np.array([]) 
                
                if len(x_data) == 0:
                    continue

                color = self._colors[orig_i % len(self._colors)]
                
                # Plot data preparation
                plot_x = []
                plot_y = []

                # Determine range to filter (Start/End)
                st = None
                en = None
                try:
                    range_mode = self.range_mode_combo.currentText() if getattr(self, 'range_mode_combo', None) else 'Default'
                    if range_mode == 'Manual' and self.start_time_spin and self.end_time_spin:
                         st = float(self.start_time_spin.value())
                         en = float(self.end_time_spin.value())
                    elif range_mode == 'Default':
                         # Default mode: do not crop data (matches get_current_range_values)
                         st = None
                         en = None
                except Exception:
                    st = en = None

                if st is not None and en is not None:
                    try:
                        arr_x = np.array(x_data)
                        arr_y = np.array(trqs)
                        mask = (arr_x >= st) & (arr_x <= en)
                        plot_x = arr_x[mask]
                        plot_y = arr_y[mask]
                    except Exception:
                        plot_x = x_data
                        plot_y = trqs
                else:
                    if sel == 'All' or sel == s['name']:
                        # If no configured range, try using spinbox values as indices or fall back to full
                         try:
                             start_idx = max(0, min(self.start_spin.value() - 1, len(trqs)-1))
                             end_idx = max(start_idx + 1, min(self.end_spin.value(), len(trqs)))
                             plot_x = x_data[start_idx:end_idx]
                             plot_y = trqs[start_idx:end_idx]
                         except:
                             plot_x = x_data
                             plot_y = trqs
                    else:
                        plot_x = x_data
                        plot_y = trqs

                # Apply K factor
                try:
                    k = self.k_spin.value() if hasattr(self, 'k_spin') else 1.0
                    if k != 1.0 and len(plot_y) > 0:
                         plot_y = [t * k for t in plot_y]
                except: pass

                if len(plot_x) > 0 and len(plot_y) > 0:
                    lw = 1.6 if sel == s['name'] else 0.9
                    alpha = 1.0 if (sel == 'All' or sel == s['name']) else 0.5
                    lbl = f"{orig_i+1}"
                    ax2.plot(plot_x, plot_y, marker='.', linestyle='-' , color=color, linewidth=lw, markersize=3, alpha=alpha, label=lbl)
                    # annotate sample number
                    try:
                        x_annot = plot_x[-1]
                        y_annot = plot_y[-1]
                        ax2.annotate(f"S{orig_i+1}", xy=(x_annot, y_annot), xytext=(6, 3), textcoords='offset points', color=color, fontsize=9, weight='bold', bbox=dict(facecolor='white', alpha=0.7, edgecolor=color, linewidth=0.5))
                    except Exception:
                        pass

        else:
            # Single dataset
            if len(self.time_data) > 0:
                try:
                    start_idx = self.start_spin.value() - 1
                    end_idx = self.end_spin.value()
                except Exception:
                    start_idx = 0
                    end_idx = len(self.time_data)
                start_idx = max(0, min(start_idx, len(self.time_data) - 1))
                end_idx = max(start_idx + 1, min(end_idx, len(self.time_data)))
                plot_time = self.time_data[start_idx:end_idx]
                plot_torque = self.torque_data[start_idx:end_idx]
                
                # Apply K factor
                try:
                    k = self.k_spin.value() if hasattr(self, 'k_spin') else 1.0
                    if k != 1.0 and len(plot_torque) > 0:
                         plot_torque = [t * k for t in plot_torque]
                except: pass

                if len(plot_time) > 0 and len(plot_torque) > 0:
                    ax2.plot(plot_time, plot_torque, 'b.-', linewidth=0.9, markersize=3, label='1')

        # Fallback: if range/cycle filters removed all lines, export the full active dataset
        # so the XLSX always contains a visible plot image.
        try:
            if len(ax2.get_lines()) == 0:
                if self.samples:
                    sel = self.file_select_combo.currentText() if hasattr(self, 'file_select_combo') else 'All'
                    sample = None
                    if sel != 'All':
                        sample = next((x for x in self.samples if x.get('name') == sel), None)
                    if sample is None:
                        sample = self.samples[0]
                    x_vals = sample.get('angle', []) if is_angle else sample.get('time', [])
                    y_vals = sample.get('torque', [])
                    if is_angle and (not x_vals or len(x_vals) != len(y_vals)):
                        x_vals = sample.get('time', [])
                    if len(x_vals) > 0 and len(y_vals) > 0:
                        ax2.plot(x_vals, y_vals, 'b.-', linewidth=0.9, markersize=3, label='1')
                elif len(self.time_data) > 0 and len(self.torque_data) > 0:
                    ax2.plot(self.time_data, self.torque_data, 'b.-', linewidth=0.9, markersize=3, label='1')
        except Exception:
            pass

        # Apply Y-axis rules to export plot
        try:
             all_y = []
             for line in ax2.get_lines():
                 y_data = line.get_ydata()
                 if y_data is not None and len(y_data) > 0:
                     all_y.extend(y_data)
             
             if all_y:
                 d_min = min(all_y)
                 d_max = max(all_y)
                 margin = (d_max - d_min) * 0.1
                 if margin == 0: margin = (d_max * 0.1) if d_max != 0 else 1.0
                 
                 t_max = d_max + margin
                 t_min = 0.0 if d_min >= 0 else d_min
                 ax2.set_ylim(t_min, t_max)
        except Exception:
             pass

        # Add legend to export figure (shows sample number labels)
        try:
            # Make legend larger and more readable in exported image
            leg = ax2.legend(loc='upper right', fontsize=11, framealpha=0.9, fancybox=True)
            leg.get_frame().set_linewidth(0.6)
        except Exception:
            pass

        try:
            fd, tmp_png = tempfile.mkstemp(suffix='.png')
            os.close(fd)
            try:
                # Reduce white margin around saved PNG: use tight bbox and configured pad.
                fig2.tight_layout(pad=IMAGE_TIGHT_PAD)
                fig2.savefig(tmp_png, format='png', dpi=DPI, bbox_inches='tight', pad_inches=IMAGE_PAD_INCHES)
            except Exception:
                fig2.savefig(tmp_png, format='png', dpi=DPI)
            # Ensure exact pixel dimensions by resizing with PIL
            try:
                from PIL import Image as PILImage
                pil = PILImage.open(tmp_png)
                pil2 = pil.resize((int(TARGET_W_PX), int(TARGET_H_PX)), PILImage.LANCZOS)
                pil2.save(tmp_png, format='PNG')
            except Exception:
                pass
        except Exception:
            if tmp_png and os.path.exists(tmp_png):
                try:
                    os.remove(tmp_png)
                except Exception:
                    pass
            tmp_png = None

        wb = Workbook()
        ws = wb.active
        ws.title = "Report"

        # ========== STYLES ==========
        bold_12 = Font(bold=True, size=12)
        bold_11 = Font(bold=True, size=11)
        title_font = Font(bold=True, size=18)
        center = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        # White fill for all sections as requested
        header_fill = PatternFill(start_color='FFFFFFFF', end_color='FFFFFFFF', fill_type='solid')
        thin_side = Side(style='thin', color='000000')
        thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        medium_side = Side(style='medium', color='000000')
        medium_border = Border(left=medium_side, right=medium_side, top=medium_side, bottom=medium_side)

        # ========== 4. TITLE ==========
        ws.merge_cells(TITLE_MERGE)
        title_cell = ws['A1']
        # Title comes from UI if available, otherwise default to 'TEST REPORT'
        try:
            title_text = self.report_title_edit.text().strip() if hasattr(self, 'report_title_edit') else ''
        except Exception:
            title_text = ''
        if not title_text:
            title_text = 'TEST REPORT'
        # Title should be uppercase, centered and bold
        title_cell.value = title_text.upper()
        title_cell.font = title_font
        title_cell.alignment = center

        # ========== 5. METADATA BLOCK (rows 3..7) and Write/Review/Approval (F1:H2) ==========
        # Layout as requested:
        # Title: A1:E2 (already merged above)
        # Write/Review/Approval: labels at F1,G1,H1 and values at F2,G2,H2
        # Left block (labels merged A:B, values C:D) rows 3..7
        # Right block (labels merged E:F, values G:H) rows 3..7
        label_fill = PatternFill(start_color='FFFFFFFF', end_color='FFFFFFFF', fill_type='solid')
        value_fill = PatternFill(start_color='FFFFFFFF', end_color='FFFFFFFF', fill_type='solid')

        # Write / Review / Approval (top-right) - labels (row 1) and values (row 2)
        wra_headers = ['Write', 'Review', 'Approval']
        wra_values = [getattr(self, 'write_edit', None).text().strip() if hasattr(self, 'write_edit') else '',
                      getattr(self, 'review_edit', None).text().strip() if hasattr(self, 'review_edit') else '',
                      getattr(self, 'approval_edit', None).text().strip() if hasattr(self, 'approval_edit') else '']
        for j, h in enumerate(wra_headers, start=6):  # F=6, G=7, H=8
            ws.cell(row=WRA_LABEL_ROW, column=j, value=h).font = bold_12
            ws.cell(row=WRA_LABEL_ROW, column=j).alignment = center
            ws.cell(row=WRA_LABEL_ROW, column=j).fill = label_fill
            ws.cell(row=WRA_LABEL_ROW, column=j).border = thin_border
            # value in row 2
            cell = ws.cell(row=WRA_VALUE_ROW, column=j, value=wra_values[j-6])
            cell.fill = value_fill
            cell.alignment = center
            cell.border = thin_border

        # Left / Right metadata rows 3..7
        rows = list(range(META_START_ROW, META_END_ROW + 1))  # 3..7
        # Prepare pairs: (left_label, left_value, right_label, right_value)
        # Build specification display from min/max spins if available
        try:
            spec_text = ''
            if getattr(self, 'spec_min_spin', None) and getattr(self, 'spec_max_spin', None):
                smin = float(self.spec_min_spin.value())
                smax = float(self.spec_max_spin.value())
                spec_text = f"{smin:.6f}-{smax:.6f} Nm"
            else:
                spec_text = getattr(self, 'spec_edit', None).text() if getattr(self, 'spec_edit', None) else ''
        except Exception:
            spec_text = ''

        pairs = [
            ('TEST ITEM', self.test_item_combo.currentText(), 'DATE', self.date_edit.date().toString('yyyy-MM-dd')),
            ('PART NAME', self.part_name_combo.currentText(), 'TESTER', self.tester_edit.text()),
            ('PART NO', self.part_no_edit.text(), 'LOT NO', self.lot_no_edit.text()),
            ('SPECIFICATION', spec_text, 'QUANTITY', str(self.quantity_spin.value())),
            ('TEST PURPOSE', (self.test_purpose_other_edit.text() if hasattr(self, 'test_purpose_combo') and self.test_purpose_combo.currentText() == "other (O)" else (self.test_purpose_combo.currentText() if hasattr(self, 'test_purpose_combo') else "")), 'JUDGMENT', self.judgment_label.text()),
            ('TEAM', self.team_combo.currentText() if hasattr(self, 'team_combo') else '', 'LINE NO', self.line_no_combo.currentText() if hasattr(self, 'line_no_combo') else '')
        ]

        for r, (l_label, l_val, r_label, r_val) in zip(rows, pairs):
            # left label merged A:B
            try:
                ws.merge_cells(f"{LEFT_LABEL_COLS[0]}{r}:{LEFT_LABEL_COLS[1]}{r}")
                cell = ws[f"{LEFT_LABEL_COLS[0]}{r}"]
                cell.value = l_label
                cell.font = bold_12
                cell.alignment = left_align
                cell.fill = label_fill
                cell.border = thin_border
            except Exception:
                pass
            # left value merged C:D
            try:
                ws.merge_cells(f"{LEFT_VALUE_COLS[0]}{r}:{LEFT_VALUE_COLS[1]}{r}")
                cell = ws[f"{LEFT_VALUE_COLS[0]}{r}"]
                cell.value = l_val
                cell.alignment = left_align
                cell.fill = value_fill
                cell.border = thin_border
            except Exception:
                pass

            # right label merged E:F
            try:
                ws.merge_cells(f"{RIGHT_LABEL_COLS[0]}{r}:{RIGHT_LABEL_COLS[1]}{r}")
                cell = ws[f"{RIGHT_LABEL_COLS[0]}{r}"]
                cell.value = r_label
                cell.font = bold_12
                cell.alignment = left_align
                cell.fill = label_fill
                cell.border = thin_border
            except Exception:
                pass
            # right value merged G:H
            try:
                ws.merge_cells(f"{RIGHT_VALUE_COLS[0]}{r}:{RIGHT_VALUE_COLS[1]}{r}")
                cell = ws[f"{RIGHT_VALUE_COLS[0]}{r}"]
                cell.value = r_val
                cell.alignment = left_align
                cell.fill = value_fill
                cell.border = thin_border
            except Exception:
                pass

        # (TORQUE TYPE and PART NAME are represented in values in the metadata table above)

        # ========== GRAPH CONDITION (Start/End Time) ==========
        # Place directly under TEST PURPOSE (row 8) and above the plot area
        try:
            graph_row = GRAPH_LABEL_ROW
            ws.cell(row=graph_row, column=1, value='GRAPH').font = bold_12
            ws.cell(row=graph_row, column=1).alignment = left_align

            # Determine start/end values based on current mode and range selection for Excel labels
            start_val = None
            end_val = None
            try:
                plot_mode_text = self.plot_mode_combo.currentText() if getattr(self, 'plot_mode_combo', None) else 'Time vs Torque'
                is_angle_mode = (plot_mode_text == "Angle vs Torque")
                range_mode = self.range_mode_combo.currentText() if getattr(self, 'range_mode_combo', None) else 'Default'
                pname = self.part_name_combo.currentText() if getattr(self, 'part_name_combo', None) else None

                # 1) Manual mode: use spinboxes
                if range_mode == 'Manual' and getattr(self, 'start_time_spin', None) and getattr(self, 'end_time_spin', None):
                    try:
                        start_val = float(self.start_time_spin.value())
                        end_val = float(self.end_time_spin.value())
                    except: pass

                # 2) Default mode: use per-part configured ranges (respecting Angle/Time mode) if they are valid
                if (start_val is None or end_val is None) and range_mode == 'Default' and pname:
                    try:
                        item_name = self.test_item_combo.currentText()
                        ranges = self.test_item_angle_ranges if is_angle_mode else self.test_item_time_ranges
                        part_ranges = ranges.get(pname, {})
                        pr = part_ranges.get(item_name, {})
                        if isinstance(pr, dict):
                            st_temp = float(pr.get('start', 0.0))
                            en_temp = float(pr.get('end', 0.0))
                            if st_temp != en_temp:
                                start_val = st_temp
                                end_val = en_temp
                    except: pass
            except: pass

            # 3) Fallback: map STT index spins to values using selected sample
            if start_val is None or end_val is None:
                try:
                    sel = self.file_select_combo.currentText() if hasattr(self, 'file_select_combo') else 'All'
                    if sel != 'All' and self.samples:
                        s = next((x for x in self.samples if x['name'] == sel), None)
                        if s:
                            data_list = s.get('angle', []) if is_angle_mode else s.get('time', [])
                            if len(data_list) == 0: data_list = s.get('time', [])
                            
                            if len(data_list) > 0:
                                si = max(0, min(self.start_spin.value() - 1, len(data_list)-1))
                                ei = max(si + 1, min(self.end_spin.value(), len(data_list)))
                                start_val = float(data_list[si])
                                end_val = float(data_list[ei-1])
                    elif self.samples and sel == 'All':
                        firsts = []
                        lasts = []
                        for s in self.samples:
                            t = s.get('angle', []) if is_angle_mode else s.get('time', [])
                            if len(t) == 0: t = s.get('time', [])
                            if len(t) > 0:
                                firsts.append(t[0])
                                lasts.append(t[-1])
                        if firsts and lasts:
                            start_val = float(min(firsts))
                            end_val = float(max(lasts))
                    else:
                        # Single dataset fallback
                        d_list = (getattr(self, 'angle_data', None) if is_angle_mode else self.time_data)
                        if d_list is None or len(d_list) == 0:
                            d_list = self.time_data
                        if len(d_list) > 0:
                            si = max(0, min(self.start_spin.value() - 1, len(d_list)-1))
                            ei = max(si + 1, min(self.end_spin.value(), len(d_list)))
                            start_val = float(d_list[si])
                            end_val = float(d_list[ei-1])
                except: pass

            # Format values for Excel; display with 6 decimal places if numeric, otherwise blank
            start_str = f"{start_val:.6f}" if start_val is not None else ''
            end_str = f"{end_val:.6f}" if end_val is not None else ''

            lbl_start = 'Start Angle (deg)' if is_angle_mode else 'Start Time (s)'
            lbl_end = 'End Angle (deg)' if is_angle_mode else 'End Time (s)'

            ws.cell(row=graph_row, column=2, value=lbl_start)
            ws.cell(row=graph_row, column=3, value=start_str)
            ws.cell(row=graph_row, column=4, value=lbl_end)
            ws.cell(row=graph_row, column=5, value=end_str)

            for cc in range(1, 6):
                ws.cell(row=graph_row, column=cc).border = thin_border
        except Exception:
            pass

        # ========== 7. GRAPH AREA ==========
        # Merge cells and add medium black border to make graph wider
        # ========== 7. GRAPH AREA ==========
        # Merge cells and add medium black border to make graph wider
        # Dynamic merge range based on calculated bottom
        graph_merge_range = f"A{GRAPH_TOP}:H{current_graph_bottom}"
        ws.merge_cells(graph_merge_range)
        for rr in range(GRAPH_TOP, current_graph_bottom + 1):
            for cc in range(1, 9):  # A=1 to H=8
                cell = ws.cell(row=rr, column=cc)
                # Clear inner borders and draw a single thin perimeter around the merged graph area
                try:
                    no_side = Side(style=None)
                except Exception:
                    no_side = Side(style=None)

                left_side = thin_side if cc == 1 else no_side
                right_side = thin_side if cc == 8 else no_side
                top_side = thin_side if rr == GRAPH_TOP else no_side
                bottom_side = thin_side if rr == current_graph_bottom else no_side

                cell.border = Border(left=left_side, right=right_side, top=top_side, bottom=bottom_side)

        # Ensure columns A..H have reasonable widths so the image fits
        try:
            ws.column_dimensions['A'].width = 16
            ws.column_dimensions['B'].width = 24
            ws.column_dimensions['C'].width = 22
            ws.column_dimensions['D'].width = 26
            # Make column E wider so "Torque (Max value)" and its values fit
            ws.column_dimensions['E'].width = 26
            ws.column_dimensions['F'].width = equal_w
            ws.column_dimensions['G'].width = equal_w
            ws.column_dimensions['H'].width = equal_w
            # Set a comfortable row height for rows GRAPH_TOP..GRAPH_BOTTOM so vertical space matches image height
            for r in range(GRAPH_TOP, current_graph_bottom + 1):
                ws.row_dimensions[r].height = 18
        except Exception:
            pass

        # Insert plot image into IMAGE_ANCHOR and scale to fit merged inner frame
        if tmp_png:
            try:
                # Compute target pixel size from worksheet column widths and row heights
                try:
                    from openpyxl.utils import get_column_letter, column_index_from_string
                except Exception:
                    get_column_letter = None
                    column_index_from_string = None

                # Determine left and right columns for the image area using the GRAPH_MERGE bounds
                try:
                    left_cell = graph_merge_range.split(':')[0]
                    right_cell = graph_merge_range.split(':')[1]
                    left_col_letters = ''.join([c for c in left_cell if c.isalpha()])
                    right_col_letters = ''.join([c for c in right_cell if c.isalpha()])
                    start_col = column_index_from_string(left_col_letters) if column_index_from_string else 1
                    end_col = column_index_from_string(right_col_letters) if column_index_from_string else 8
                except Exception:
                    start_col = 1
                    end_col = 8

                # Rows from IMAGE_TOP..IMAGE_BOTTOM (dynamic)
                top_row = IMAGE_TOP
                bottom_row = current_image_bottom

                # Calculate total width in pixels across columns start_col..end_col
                total_w_px = 0
                for c in range(start_col, end_col + 1):
                    try:
                        col_letter = get_column_letter(c) if get_column_letter else 'A'
                        col_dim = ws.column_dimensions.get(col_letter)
                        col_w = col_dim.width if col_dim and col_dim.width is not None else 8.43
                        # Approx conversion: pixels = width*7 + 5
                        px = int(col_w * 7 + 5)
                    except Exception:
                        px = 64
                    total_w_px += px

                # Calculate total height in pixels across rows top_row..bottom_row
                total_h_px = 0
                for r in range(top_row, bottom_row + 1):
                    try:
                        row_dim = ws.row_dimensions.get(r)
                        row_h = row_dim.height if row_dim and row_dim.height is not None else 15
                        # Convert points to pixels: px = points * 96 / 72
                        px = int(row_h * 96.0 / 72.0)
                    except Exception:
                        px = 18
                    total_h_px += px

                # Compensate for Excel internal margins and rendering quirks by expanding
                # the target a few pixels so the image visually fills the merged frame.
                pad_w = 0
                pad_h = 0
                # Also add the width of the right-most column to ensure the image
                # reaches the final merged column (helps when image was stopping at G)
                try:
                    last_col_letter = get_column_letter(end_col) if get_column_letter else None
                    last_col_dim = ws.column_dimensions.get(last_col_letter) if last_col_letter else None
                    last_col_w = last_col_dim.width if last_col_dim and last_col_dim.width is not None else 8.43
                    last_col_px = int(last_col_w * 7 + 5)
                except Exception:
                    last_col_px = 0

                target_w = max(1, int(total_w_px - pad_w + IMAGE_EXPAND_PX_W + last_col_px))
                target_h = max(1, int(total_h_px - pad_h + IMAGE_EXPAND_PX_H))

                # Render at higher resolution and downsample to make the image sharper
                try:
                    pil = PILImage.open(tmp_png)
                    src_w, src_h = pil.size
                    # Render scale: produce a larger source then downsample for sharpness
                    render_w = min(int(target_w * IMAGE_RENDER_SCALE), IMAGE_MAX_RENDER_PX)
                    render_h = min(int(target_h * IMAGE_RENDER_SCALE), IMAGE_MAX_RENDER_PX)

                    try:
                        # Always re-render at the higher resolution (gives consistent quality)
                        try:
                            fig2.set_size_inches(max(fig2.get_size_inches()[0], render_w / DPI),
                                                 max(fig2.get_size_inches()[1], render_h / DPI))
                            fig2.tight_layout(pad=IMAGE_TIGHT_PAD)
                            fig2.savefig(tmp_png, format='png', dpi=DPI, bbox_inches='tight', pad_inches=IMAGE_PAD_INCHES)
                        except Exception:
                            fig2.savefig(tmp_png, format='png', dpi=DPI)
                        pil = PILImage.open(tmp_png)
                        src_w, src_h = pil.size
                    except Exception:
                        pass

                    # Finally downsample to exact target using high-quality resampling
                    try:
                        pil2 = pil.resize((target_w, target_h), PILImage.LANCZOS)
                        pil2.save(tmp_png, format='PNG')
                    except Exception:
                        # fallback: keep original
                        pass
                except Exception:
                    pass

                try:
                    img = XLImage(tmp_png)
                    try:
                        img.width = target_w
                        img.height = target_h
                    except Exception:
                        pass
                    # Anchor image at the left column of the GRAPH merge and one row
                    # below GRAPH_TOP so it sits inside the merged frame correctly.
                    try:
                        # Compute a half-column offset (in pixels) so the image is nudged
                        # slightly right (about half of the left column width).
                        try:
                            col_dim0 = ws.column_dimensions.get(left_col_letters)
                            col_w0 = col_dim0.width if col_dim0 and col_dim0.width is not None else 8.43
                            col_px0 = int(col_w0 * 7 + 5)
                        except Exception:
                            col_px0 = 8

                        EMU_PER_PIXEL = IMAGE_EMU_PER_PIXEL
                        try:
                            # Support integer column shifts plus a fractional remainder.
                            # e.g. IMAGE_HALF_COL_OFFSET_FRAC = 1.5 -> shift 1 whole
                            # column and then 0.5 of the next column.
                            frac = float(IMAGE_HALF_COL_OFFSET_FRAC)
                            whole_cols = int(frac)
                            remainder = max(0.0, frac - whole_cols)

                            # Determine anchor column after whole-column shift
                            anchor_col_index = min(end_col, start_col + whole_cols)

                            if remainder > 0:
                                # Compute fractional pixel offset based on the left-most column width
                                frac_px = int(col_px0 * remainder)
                                # Some environments ignore positive colOff. A more
                                # compatible approach is to anchor at the next
                                # column and apply a negative colOff so the image
                                # shifts left into the previous column by
                                # (col_px0 - frac_px) pixels (i.e. the requested
                                # fractional gap).
                                try:
                                    # anchor one column to the right and move left
                                    anchor_for_neg = min(end_col, anchor_col_index + 1)
                                    img.anchor._from.col = anchor_for_neg - 1
                                    img.anchor._from.row = top_row - 1
                                    neg_px = int(col_px0 - frac_px)
                                    img.anchor._from.colOff = -int(neg_px * EMU_PER_PIXEL)
                                    img.anchor._from.rowOff = 0
                                    ws.add_image(img)
                                except Exception:
                                    # Fallback to positive colOff at anchor_col_index
                                    try:
                                        img.anchor._from.col = anchor_col_index - 1
                                        img.anchor._from.row = top_row - 1
                                        img.anchor._from.colOff = int(frac_px * EMU_PER_PIXEL)
                                        img.anchor._from.rowOff = 0
                                        ws.add_image(img)
                                    except Exception:
                                        try:
                                            anchor_cell = f"{get_column_letter(anchor_col_index)}{top_row}"
                                        except Exception:
                                            anchor_cell = IMAGE_ANCHOR
                                        ws.add_image(img, anchor_cell)
                            else:
                                # No fractional remainder, anchor at the whole-column position
                                try:
                                    anchor_cell = f"{get_column_letter(anchor_col_index)}{top_row}"
                                except Exception:
                                    anchor_cell = IMAGE_ANCHOR
                                ws.add_image(img, anchor_cell)
                        except Exception:
                            # Final fallback: place at GRAPH_TOP+1 with no offset
                            try:
                                anchor_col_index = min(end_col, start_col)
                                anchor_col_letter = get_column_letter(anchor_col_index) if get_column_letter else left_col_letters
                                anchor_cell = f"{anchor_col_letter}{GRAPH_TOP + 1}"
                            except Exception:
                                anchor_cell = IMAGE_ANCHOR
                            try:
                                ws.add_image(img, anchor_cell)
                            except Exception:
                                try:
                                    ws.add_image(img)
                                except Exception:
                                    pass
                    except Exception:
                        try:
                            ws.add_image(img, IMAGE_ANCHOR)
                        except Exception:
                            pass
                except Exception:
                    pass
            except Exception:
                try:
                    img = XLImage(tmp_png)
                    ws.add_image(img, IMAGE_ANCHOR)
                except Exception:
                    pass

        # Write/Review/Approval handled above at rows 1..2

        # ========== 8. SAMPLE SUMMARY TABLE (below image) ==========
        # Move sample table further down to avoid being overlapped by the image
        # ========== 8. SAMPLE SUMMARY TABLE (below image) ==========
        # Move sample table further down to avoid being overlapped by the image
        sample_header_row = current_graph_bottom + 1
        sample_headers = ['Sample No.', 'PART NO', 'Average Torque (Nm)', 'Average Torque (Kgf.cm)', 'Max Torque (Nm)', 'Max Torque (Kgf.cm)', 'Min Torque (Nm)', 'Min Torque (Kgf.cm)']
        for c, h in enumerate(sample_headers, start=1):
            cell = ws.cell(row=sample_header_row, column=c, value=h)
            cell.font = bold_11
            cell.alignment = center
            # cell.fill = header_fill  # Removed yellow color
            cell.border = thin_border

        # Decide which samples to write: include all imported samples
        if self.samples:
            samples_to_write = list(self.samples)
        else:
            # single dataset fallback: respect PART NO from UI (may be empty)
            samples_to_write = [{
                'path': self.file_path or '',
                'name': self.file_label.text(),
                'time': self.time_data,
                'torque': self.torque_data,
                'part_no': self.part_no_edit.text().strip() or ''
            }]

        data_row = sample_header_row + 1
        for idx, sample in enumerate(samples_to_write, start=1):
            ws.cell(row=data_row, column=1, value=idx).alignment = center
            # PART NO should come from the sample if recorded at import time;
            # otherwise fall back to the current PART NO field in the UI so
            # exported table always shows the application's PART NO value.
            part_no_val = sample.get('part_no') or (self.part_no_edit.text().strip() if hasattr(self, 'part_no_edit') else '')
            ws.cell(row=data_row, column=2, value=part_no_val)

            # Use shared helper logic to get values for this sample based on exact current range request
            vals = self.get_current_range_values(sample)
            
            if len(vals) > 0:
                import numpy as np
                # Convert to absolute values to match App calculation logic
                vals_abs = np.abs(vals)
                ravg = float(np.mean(vals_abs))
                rmin = float(np.min(vals))
                rmax = float(np.max(vals))
                # Kgf.cm (1 Nm = 10.1972 Kgf.cm)
                ravg_k = ravg * 10.1972
                rmin_k = rmin * 10.1972
                rmax_k = rmax * 10.1972
            else:
                ravg = rmin = rmax = ravg_k = rmin_k = rmax_k = None

            # Write stats with high numeric precision format
            # Columns: 1=Sample, 2=PartNo, 3=Avg(Nm), 4=Avg(Kgf.cm), 5=Max(Nm), 6=Max(Kgf.cm), 7=Min(Nm), 8=Min(Kgf.cm)
            c3 = ws.cell(row=data_row, column=3, value=ravg)
            c4 = ws.cell(row=data_row, column=4, value=ravg_k)
            c5 = ws.cell(row=data_row, column=5, value=rmax)
            c6 = ws.cell(row=data_row, column=6, value=rmax_k)
            c7 = ws.cell(row=data_row, column=7, value=rmin)
            c8 = ws.cell(row=data_row, column=8, value=rmin_k)

            try:
                fmt = '0.00000000'
                for cell in [c3, c4, c5, c6, c7, c8]:
                    if cell.value is not None:
                        cell.number_format = fmt
            except Exception:
                pass
            
            # Apply borders to all cells in this row
            for c in range(1, 9):
                ws.cell(row=data_row, column=c).border = thin_border
            data_row += 1
            
        # Adjust row heights 1 and 2 and ensure E2 alignment as requested - set to 40 to match other metadata rows
        try:
            ws.row_dimensions[1].height = 40
            ws.row_dimensions[2].height = 40
        except Exception:
            pass

        # ========== OUTER BORDER AROUND MAIN REPORT AREA ==========
        try:
            # Draw a medium black border around columns A..H and rows 1..(data_row+1)
            outer_start_row = 1
            outer_end_row = max(data_row + 1, current_graph_bottom + 1)
            for rr in range(outer_start_row, outer_end_row + 1):
                for cc in range(1, 9):
                    cell = ws.cell(row=rr, column=cc)
                    # Determine which sides are the outer perimeter
                    left = medium_side if cc == 1 else thin_side
                    right = medium_side if cc == 8 else thin_side
                    top = medium_side if rr == outer_start_row else thin_side
                    bottom = medium_side if rr == outer_end_row else thin_side
                    cell.border = Border(left=left, right=right, top=top, bottom=bottom)
        except Exception:
            pass

        # Note: GRAPH condition already placed above the plot (row 9)

        # ========== 10. SHEET "DATA" ==========
        data_sheet = wb.create_sheet('Data')
        # Detect mode robustly (case-insensitive, allow variations)
        mode_text = (self.plot_mode_combo.currentText() or '').strip().lower() if getattr(self, 'plot_mode_combo', None) else ''
        is_angle = mode_text.startswith('angle')
        x_label = 'Angle (deg)' if is_angle else 'Time (s)'
        
        if self.samples:
            data_sheet.cell(row=1, column=1, value='File').font = bold_12
            data_sheet.cell(row=1, column=2, value='Index').font = bold_12
            data_sheet.cell(row=1, column=3, value=x_label).font = bold_12
            data_sheet.cell(row=1, column=4, value='Torque (N·m)').font = bold_12
            row = 2
            for s in samples_to_write:
                name = s.get('name') or os.path.basename(s.get('path', '') or '')
                # Choose X-axis data based on mode with better fallbacks and length handling
                if is_angle:
                    x_data = s.get('angle', None)
                    trqs = s.get('torque', []) or []
                    if x_data is None:
                        # no angle array available -> fallback to time
                        x_data = s.get('time', []) or []
                    else:
                        # if angle length mismatches torque, try to align by truncating or padding
                        try:
                            if len(x_data) != len(trqs):
                                # if longer, truncate; if shorter, pad with last value or fallback to time
                                if len(x_data) > len(trqs):
                                    x_data = x_data[:len(trqs)]
                                elif len(x_data) < len(trqs):
                                    # fallback to time if available
                                    tdata = s.get('time', []) or []
                                    if len(tdata) == len(trqs):
                                        x_data = tdata
                                    else:
                                        # pad angle with last known value
                                        last = x_data[-1] if len(x_data) > 0 else 0
                                        x_data = list(x_data) + [last] * (len(trqs) - len(x_data))
                        except Exception:
                            x_data = s.get('time', []) or []
                else:
                    x_data = s.get('time', []) or []

                trqs = s.get('torque', []) or []
                
                for i, (x_val, trq) in enumerate(zip(x_data, trqs), start=1):
                    data_sheet.cell(row=row, column=1, value=name)
                    data_sheet.cell(row=row, column=2, value=i)
                    ct = data_sheet.cell(row=row, column=3, value=x_val)
                    ctr = data_sheet.cell(row=row, column=4, value=trq)
                    try:
                        ct.number_format = '0.00000000'
                        ctr.number_format = '0.00000000'
                    except Exception:
                        pass
                    row += 1
        else:
            # single dataset stored in self.time_data/self.torque_data
            data_sheet.cell(row=1, column=1, value='Index').font = bold_12
            data_sheet.cell(row=1, column=2, value=x_label).font = bold_12
            data_sheet.cell(row=1, column=3, value='Torque (N·m)').font = bold_12
            
            # For single dataset, we might not have 'angle_data' separate if not in samples list?
            # Actually we usually store it in self.samples. 
            # If self.samples is empty, self.time_data is used.
            # But where is single dataset angle stored?
            # If loaded via import_csv, it goes to samples.
            # If loaded differently? The app seems to always use self.samples now after import.
            # But let's check self.time_data fallback.
            # If we only have self.time_data, we might not have angle.
            # Assuming Single Dataset -> Time vs Torque usually.
            
            # If in Angle Mode but only Time Data available?
            # Try to use stored angle array if present in the object; otherwise fallback to time_data
            if is_angle and getattr(self, 'angle_data', None):
                x_data = getattr(self, 'angle_data')
            else:
                x_data = self.time_data
            # If we don't have angle data for single dataset fallback, we can't do much.
            # But commonly self.samples is populated. This fallback is for safety.
            
            for i, (x_val, trq) in enumerate(zip(x_data, self.torque_data), start=1):
                data_sheet.cell(row=i + 1, column=1, value=i)
                ct = data_sheet.cell(row=i + 1, column=2, value=x_val)
                ctr = data_sheet.cell(row=i + 1, column=3, value=trq)
                try:
                    ct.number_format = '0.00000000'
                    ctr.number_format = '0.00000000'
                except Exception:
                    pass

        # ========== COLUMN WIDTHS ==========
        ws.column_dimensions['A'].width = 16
        ws.column_dimensions['B'].width = 24
        ws.column_dimensions['C'].width = 22
        ws.column_dimensions['D'].width = 26
        # Ensure E is wide enough for the sample summary (Torque columns)
        # and set F/G/H equal width for the report area to the right
        # Make column E wider so sample torque max values have enough room
        ws.column_dimensions['E'].width = 26
        equal_w = 18
        ws.column_dimensions['F'].width = equal_w
        ws.column_dimensions['G'].width = equal_w
        ws.column_dimensions['H'].width = equal_w
        ws.column_dimensions['I'].width = 12
        ws.column_dimensions['J'].width = 12
        ws.column_dimensions['K'].width = 12
        ws.column_dimensions['L'].width = 12

        # Increase row heights for metadata rows (3..7) so all metadata fields
        # (TEST ITEM, PART NAME, PART NO, SPECIFICATION, TEST PURPOSE) have equal height
        try:
            for r in range(META_START_ROW, META_END_ROW + 1):
                ws.row_dimensions[r].height = 40
        except Exception:
            pass

        # ========== A4 PAGE SETUP ==========
        try:
            ws.page_setup.paperSize = 9  # A4
            ws.page_setup.orientation = 'portrait'
            ws.page_setup.fitToPage = True
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0
        except Exception:
            pass

        # ========== SAVE ==========
        try:
            wb.save(path)
            QMessageBox.information(self, "Saved", f"Excel report saved to:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to save Excel:\n{e}")
        finally:
            if tmp_png and os.path.exists(tmp_png):
                try:
                    os.remove(tmp_png)
                except Exception:
                    pass
    def clear_remark(self):
        """Clear one-time Remark when leaving the analysis/report workflow."""
        try:
            if hasattr(self, 'remark_edit'):
                self.remark_edit.clear()
        except Exception:
            pass

    def save_profile(self):
        """Save current metadata as a JSON profile."""
        suggested = 'profile.json'
        path, _ = QFileDialog.getSaveFileName(self, "Save Profile", suggested, "JSON Files (*.json);;All Files (*)")
        if not path:
            return
        profile = {
            'test_item': self.test_item_combo.currentText(),
            'part_name': self.part_name_combo.currentText(),
            'part_no': self.part_no_edit.text().strip(),
            'report_title': self.report_title_edit.text().strip() if hasattr(self, 'report_title_edit') else 'TEST REPORT',
            'date': self.date_edit.date().toString('yyyy-MM-dd'),
            'tester': self.tester_edit.text().strip(),
            'write': self.write_edit.text().strip() if hasattr(self, 'write_edit') else '',
            'review': self.review_edit.text().strip() if hasattr(self, 'review_edit') else '',
            'approval': self.approval_edit.text().strip() if hasattr(self, 'approval_edit') else '',
            'test_purpose': (self.test_purpose_other_edit.text() if hasattr(self, 'test_purpose_combo') and self.test_purpose_combo.currentText() == "other (O)" else (self.test_purpose_combo.currentText() if hasattr(self, 'test_purpose_combo') else "")),
            'lot_no': self.lot_no_edit.text().strip() if hasattr(self, 'lot_no_edit') else '',
            'quantity': int(self.quantity_spin.value()) if hasattr(self, 'quantity_spin') else 1,
            'stt_start': int(self.start_spin.value()) if hasattr(self, 'start_spin') else 1,
            'stt_end': int(self.end_spin.value()) if hasattr(self, 'end_spin') else 100,
            'team': self.team_combo.currentText() if hasattr(self, 'team_combo') else '',
            'line_no': self.line_no_combo.currentText() if hasattr(self, 'line_no_combo') else '',
            'sample_no': int(self.sample_no_spin.value()) if hasattr(self, 'sample_no_spin') else 1,
        }
        try:
            if getattr(self, 'spec_min_spin', None) and getattr(self, 'spec_max_spin', None):
                smin = float(self.spec_min_spin.value())
                smax = float(self.spec_max_spin.value())
                profile['spec_min'] = smin
                profile['spec_max'] = smax
                profile['spec'] = f"Min: {smin:.6f} Nm  Max: {smax:.6f} Nm"
            elif getattr(self, 'spec_edit', None):
                profile['spec'] = self.spec_edit.text().strip()
        except Exception:
            try:
                if getattr(self, 'spec_edit', None):
                    profile['spec'] = self.spec_edit.text().strip()
            except Exception:
                pass
        try:
            with open(path, 'w', encoding='utf-8') as f:
                json.dump(profile, f, ensure_ascii=False, indent=2)
            QMessageBox.information(self, "Profile Saved", f"Profile saved to:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to save profile:\n{e}")

    def load_profile(self):
        """Load metadata from a JSON profile and populate the form."""
        path, _ = QFileDialog.getOpenFileName(self, "Load Profile", "", "JSON Files (*.json);;All Files (*)")
        if not path:
            return
        try:
            with open(path, 'r', encoding='utf-8') as f:
                profile = json.load(f)
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to load profile:\n{e}")
            return

        try:
            if 'test_item' in profile:
                try:
                    self.test_item_combo.setCurrentText(profile.get('test_item', ''))
                except Exception:
                    pass
            if 'report_title' in profile and hasattr(self, 'report_title_edit'):
                try:
                    self.report_title_edit.setText(profile.get('report_title', 'TEST REPORT'))
                except Exception:
                    pass
            if 'part_name' in profile:
                try:
                    self.part_name_combo.setCurrentText(profile.get('part_name', ''))
                except Exception:
                    pass
            if 'part_no' in profile:
                self.part_no_edit.setText(profile.get('part_no', ''))
            if 'sample_no' in profile and hasattr(self, 'sample_no_spin'):
                try:
                    self.sample_no_spin.setValue(max(1, min(99, int(profile.get('sample_no', 1)))))
                except Exception:
                    pass
            # Apply spec: prefer numeric spec_min/spec_max if present, otherwise try textual spec
            if 'spec_min' in profile and 'spec_max' in profile and getattr(self, 'spec_min_spin', None) and getattr(self, 'spec_max_spin', None):
                try:
                    self.spec_min_spin.setValue(float(profile.get('spec_min', 0.0)))
                    self.spec_max_spin.setValue(float(profile.get('spec_max', 0.0)))
                except Exception:
                    pass
            elif 'spec' in profile:
                try:
                    if getattr(self, 'spec_min_spin', None) and getattr(self, 'spec_max_spin', None):
                        import re
                        s = profile.get('spec', '')
                        nums = re.findall(r"[-+]?[0-9]*\.?[0-9]+", s)
                        if len(nums) >= 2:
                            try:
                                self.spec_min_spin.setValue(float(nums[0]))
                                self.spec_max_spin.setValue(float(nums[1]))
                            except Exception:
                                pass
                    else:
                        if getattr(self, 'spec_edit', None):
                            self.spec_edit.setText(profile.get('spec', ''))
                except Exception:
                    try:
                        if getattr(self, 'spec_edit', None):
                            self.spec_edit.setText(profile.get('spec', ''))
                    except Exception:
                        pass
            if 'date' in profile:
                try:
                    d = QDate.fromString(profile.get('date', ''), 'yyyy-MM-dd')
                    if d.isValid():
                        self.date_edit.setDate(d)
                except Exception:
                    pass
            if 'tester' in profile:
                self.tester_edit.setText(profile.get('tester', ''))
            if 'write' in profile and hasattr(self, 'write_edit'):
                self.write_edit.setText(profile.get('write', ''))
            if 'review' in profile and hasattr(self, 'review_edit'):
                self.review_edit.setText(profile.get('review', ''))
            if 'approval' in profile and hasattr(self, 'approval_edit'):
                self.approval_edit.setText(profile.get('approval', ''))
            if 'team' in profile and hasattr(self, 'team_combo'):
                self.team_combo.setCurrentText(profile.get('team', ''))
            if 'line_no' in profile and hasattr(self, 'line_no_combo'):
                self.line_no_combo.setCurrentText(profile.get('line_no', ''))
            if 'test_purpose' in profile and hasattr(self, 'test_purpose_combo'):
                val = profile.get('test_purpose', '')
                if val:
                    self.test_purpose_combo.blockSignals(True)
                    try:
                        idx = self.test_purpose_combo.findText(val)
                        if idx >= 0 and val != "other (O)":
                             self.test_purpose_combo.setCurrentIndex(idx)
                             self.test_purpose_other_edit.hide()
                        else:
                             idx_other = self.test_purpose_combo.findText("other (O)")
                             if idx_other >= 0:
                                self.test_purpose_combo.setCurrentIndex(idx_other)
                                self.test_purpose_other_edit.setText(val)
                                self.test_purpose_other_edit.show()
                    finally:
                        self.test_purpose_combo.blockSignals(False)
                else:
                    self.test_purpose_combo.setCurrentIndex(0)
                    self.test_purpose_other_edit.clear()
                    self.test_purpose_other_edit.hide()
            if 'lot_no' in profile and hasattr(self, 'lot_no_edit'):
                self.lot_no_edit.setText(profile.get('lot_no', ''))
            if 'quantity' in profile and hasattr(self, 'quantity_spin'):
                try:
                    self.quantity_spin.setValue(int(profile.get('quantity', 1)))
                except Exception:
                    pass
            if 'stt_start' in profile and hasattr(self, 'start_spin'):
                try:
                    self.start_spin.setValue(int(profile.get('stt_start', 1)))
                except Exception:
                    pass
            if 'stt_end' in profile and hasattr(self, 'end_spin'):
                try:
                    self.end_spin.setValue(int(profile.get('stt_end', 1)))
                except Exception:
                    pass

            QMessageBox.information(self, "Profile Loaded", f"Profile loaded from:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to apply profile:\n{e}")

    def browse_csv_path(self):
        dir_path = QFileDialog.getExistingDirectory(self, "Select CSV Save Directory", self.csv_path_edit.text())
        if dir_path:
            self.csv_path_edit.setText(dir_path)
            
    def browse_report_path(self):
        dir_path = QFileDialog.getExistingDirectory(self, "Select Report Save Directory", self.report_path_edit.text())
        if dir_path:
            self.report_path_edit.setText(dir_path)

    def save_report(self):
        """Save both raw CSV and CTR report CSV using ReportService."""
        if not self.samples:
            QMessageBox.warning(self, "No Data", "No imported samples to save.")
            return

        sel = self.file_select_combo.currentText()
        if sel == "All":
            active_sample = self.samples[0]
        else:
            active_sample = next((x for x in self.samples if x['name'] == sel), None)

        if not active_sample:
            QMessageBox.warning(self, "No Data", "Active sample not found.")
            return

        csv_dir = self.csv_path_edit.text().strip()
        report_dir = self.report_path_edit.text().strip()

        if not csv_dir or not report_dir:
            QMessageBox.warning(self, "Missing Path", "Please select both CSV File path and Report File path directories.")
            return

        if not os.path.exists(csv_dir):
            try:
                os.makedirs(csv_dir, exist_ok=True)
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Could not create CSV directory:\n{e}")
                return

        if not os.path.exists(report_dir):
            try:
                os.makedirs(report_dir, exist_ok=True)
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Could not create Report directory:\n{e}")
                return

        from domain.entities import RecordingSession, SampleData, ReportMetadata

        # 1. Create list of SampleData from active_sample
        times = active_sample.get('time', [])
        torques = active_sample.get('torque', [])
        angles = active_sample.get('angle', [])
        cycles = active_sample.get('cycle', [])

        # Determine average sample interval in ms
        interval_ms = 100
        if len(times) > 1:
            try:
                interval_ms = int(round((times[1] - times[0]) * 1000.0))
                if interval_ms <= 0:
                    interval_ms = 100
            except:
                pass

        k = getattr(self, 'k_factor', 1.0)
        raw_samples = []
        for i in range(len(times)):
            t = times[i]
            trq = torques[i]
            if k != 1.0:
                trq = trq * k
            ang = angles[i] if i < len(angles) else 0.0
            cyc = cycles[i] if i < len(cycles) else 1
            raw_samples.append(SampleData(
                time_s=t,
                torque_Nm=trq,
                stable=True,
                angle_deg=ang,
                cycle=cyc
            ))

        raw_session = RecordingSession(
            samples=raw_samples,
            sample_interval_ms=interval_ms
        )

        # 2. Get trimmed samples for CTR report
        # Determine whether current plot mode is Angle or Time
        try:
            mode_text = (self.plot_mode_combo.currentText() or '').strip().lower() if getattr(self, 'plot_mode_combo', None) else ''
            is_angle_mode = mode_text.startswith('angle')
        except Exception:
            is_angle_mode = False

        range_mode = self.range_mode_combo.currentText() if getattr(self, 'range_mode_combo', None) else 'Default'
        st = None
        en = None

        if range_mode == 'Manual':
            if getattr(self, 'start_time_spin', None) and getattr(self, 'end_time_spin', None):
                st = float(self.start_time_spin.value())
                en = float(self.end_time_spin.value())
        else:
            pname = self.part_name_combo.currentText() if getattr(self, 'part_name_combo', None) else None
            item_name = self.test_item_combo.currentText()
            ranges = self.test_item_angle_ranges if is_angle_mode else self.test_item_time_ranges
            part_ranges = ranges.get(pname, {})
            pr = part_ranges.get(item_name, {})
            if isinstance(pr, dict):
                st = float(pr.get('start', 0.0))
                en = float(pr.get('end', 0.0))

        selected_c = active_sample.get('selected_cycles', [])
        trimmed_samples = []
        for s in raw_samples:
            # 1. Cycle selection check
            if selected_c and s.cycle not in selected_c:
                continue
            # 2. Bound check
            if st is not None and en is not None:
                x_val = s.angle_deg if is_angle_mode else s.time_s
                if x_val < st or x_val > en:
                    continue
            trimmed_samples.append(s)

        if not trimmed_samples:
            trimmed_samples = raw_samples

        # Create ReportMetadata
        metadata = ReportMetadata(
            test_item=self.test_item_combo.currentText(),
            part_name=self.part_name_combo.currentText(),
            part_no=self.part_no_edit.text().strip(),
            sample_no=int(self.sample_no_spin.value()) if hasattr(self, 'sample_no_spin') else 1,
            remark=self.remark_edit.text().strip() if hasattr(self, 'remark_edit') else '',
            test_purpose=self.test_purpose_combo.currentText(),
            tester=self.tester_edit.text().strip(),
            team=self.team_combo.currentText() if hasattr(self, 'team_combo') else '',
            line_no=self.line_no_combo.currentText() if hasattr(self, 'line_no_combo') else '',
            date=self.date_edit.date().toString('yyyy-MM-dd'),
            csv_path=csv_dir,
            report_path=report_dir
        )

        from application.report_service import ReportService
        report_svc = ReportService()

        try:
            filename = report_svc.generate_filename(metadata, csv_dir)
            raw_path = report_svc.save_raw_csv(raw_session, csv_dir, filename)
            report_path = report_svc.save_ctr_report(
                trimmed_samples=trimmed_samples,
                report_dir=report_dir,
                filename=filename,
                session_interval_ms=interval_ms,
                metadata=metadata
            )

            QMessageBox.information(
                self, "Report Saved",
                f"Successfully saved both reports:\n\n1. Raw CSV: {raw_path}\n2. CTR Report: {report_path}"
            )

            # Save report paths to report_paths.json
            self.csv_dir = csv_dir
            self.report_dir = report_dir
            try:
                cfg = get_config_file('report_paths.json')
                with open(cfg, 'w', encoding='utf-8') as f:
                    json.dump({'csv_dir': csv_dir, 'report_dir': report_dir}, f, indent=2)
            except:
                pass

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to save reports:\n{e}")

def main():
    app = QApplication(sys.argv)
    window = TorquePlotViewer()
    window.show()
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()
